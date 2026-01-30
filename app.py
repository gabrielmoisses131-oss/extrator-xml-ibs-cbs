# -*- coding: utf-8 -*-
"""
Extrator XML -> Planilha (IBS/CBS)
- Faz upload de 1+ XML (NFe/NFCe) e opcionalmente uma planilha modelo (.xlsx)
- Extrai: Data, Número da Nota, Item/Serviço, cClassTrib, Base (vBC), vIBS, vCBS, arquivo, Fonte do valor
- Grava na aba "LANCAMENTOS" preservando fórmulas/validações existentes (Excel recalcula ao abrir)

Como rodar:
  python -m pip install -r requirements.txt
  python -m streamlit run app.py
"""
import io
import zipfile
from datetime import datetime, date
import xml.etree.ElementTree as ET

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
import html
import time
import hashlib
from openpyxl import load_workbook
from textwrap import dedent

# -----------------------------
# Page config + CSS (Figma-like)
# -----------------------------
st.set_page_config(page_title="Extrator XML - IBS/CBS", layout="wide")

CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

:root{
  --card: rgba(255,255,255,.92);
  --card2: rgba(255,255,255,.82);
  --ink: #0f172a;
  --muted:#64748b;
  --line: rgba(15,23,42,.10);
  --shadow: 0 18px 45px rgba(2,6,23,.10);
  --shadow2: 0 26px 70px rgba(2,6,23,.16);
  --radius: 18px;

  --blue:#2563eb;
  --green:#16a34a;
  --amber:#f59e0b;
  --purple:#7c3aed;
}

.stApp{
  background:
    radial-gradient(1200px 520px at 12% -10%, rgba(37,99,235,.18), transparent 45%),
    radial-gradient(900px 520px at 110% 10%, rgba(124,58,237,.18), transparent 50%),
    radial-gradient(900px 520px at 40% 120%, rgba(22,163,74,.14), transparent 45%),
    #f6f8fc !important;
  font-family: Inter, system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif !important;
}

.block-container{
  padding-top: 1.5rem !important;
  padding-bottom: 2.5rem !important;
  max-width: 1200px !important;
}

h1,h2,h3,h4,h5,h6{ color: var(--ink) !important; letter-spacing: -.02em; }
p,li,span,small,.stCaption{ color: var(--muted) !important; }

/* Sidebar (coluna arredondada estilo app) */
section[data-testid="stSidebar"]{
  background: radial-gradient(800px 520px at 30% 0%, rgba(37,99,235,.35), transparent 50%),
              radial-gradient(800px 520px at 70% 80%, rgba(124,58,237,.35), transparent 55%),
              #0b1220 !important;

  /* Contorno + formato de “painel” */
  border: 1px solid rgba(255,255,255,.12) !important;
  border-radius: 22px !important;
  box-shadow: 0 20px 55px rgba(0,0,0,.35), inset 0 0 0 1px rgba(255,255,255,.05) !important;

  /* Respiro para parecer coluna flutuante */
  margin: 12px !important;
  overflow: hidden !important;
}

/* Garante que o conteúdo interno respeite o arredondado e ocupe a altura toda */
section[data-testid="stSidebar"] > div{
  border-radius: 22px !important;
  overflow: hidden !important;
  height: calc(100vh - 24px) !important;
}

/* Texto/cores dentro da sidebar */
section[data-testid="stSidebar"] *{ color: rgba(255,255,255,.92) !important; }
section[data-testid="stSidebar"] .stCaption,
section[data-testid="stSidebar"] small{ color: rgba(255,255,255,.65) !important; }

/* Cards */
.card{
  background: linear-gradient(180deg, var(--card), var(--card2));
  border: 1px solid var(--line);
  border-radius: var(--radius);
  box-shadow: var(--shadow);
  padding: 18px 20px;
  backdrop-filter: blur(10px);
}
.card + .card{ margin-top: 14px; }

/* Top header (match premium mock) */
.topbar{
  background: linear-gradient(180deg, rgba(255,255,255,.96), rgba(255,255,255,.86));
  border: 1px solid rgba(15,23,42,.10);
  border-radius: 18px;
  box-shadow: var(--shadow);
  padding: 16px 18px;
  display:flex;
  align-items:center;
  justify-content:space-between;
  gap: 16px;
}
.topbar .brand{ display:flex; align-items:center; gap: 12px; }
.brand-badge{
  width: 44px;
  height: 44px;
  border-radius: 16px;
  background: rgba(15,23,42,.04);
  border: 1px solid rgba(15,23,42,.08);
  display:flex;
  align-items:center;
  justify-content:center;
  flex: 0 0 auto;
}
.brand-badge svg{ width: 22px; height: 22px; }
.topbar h1{ margin:0; font-size: 1.45rem; font-weight: 900; color:#0f172a; }
.topbar .sub{ margin-top:2px; font-size:.92rem; color:#64748b; }
.status-pill{
  display:inline-flex;
  align-items:center;
  gap: 8px;
  padding: 8px 12px;
  border-radius: 999px;
  background: rgba(15,23,42,.04);
  border: 1px solid rgba(15,23,42,.08);
  font-weight: 800;
  color:#475569;
  white-space: nowrap;
}
.status-dot{ width:8px; height:8px; border-radius: 999px; background: #64748b; }

.hr{ height:1px; background: rgba(15,23,42,.10); margin: 18px 0; }

.pill{
  display:inline-flex;
  align-items:center;
  gap:8px;
  padding: 7px 11px;
  border-radius: 999px;
  background: rgba(15,23,42,.05);
  border: 1px solid rgba(15,23,42,.08);
  color: var(--muted);
  font-weight: 700;
  font-size: .82rem;
}

/* KPI grid + clickable cards */
.kpi-grid{ display:grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 16px; }
@media(max-width:1200px){ .kpi-grid{ grid-template-columns: repeat(2, 1fr);} }
@media(max-width:650px){ .kpi-grid{ grid-template-columns: 1fr;} }

.kpi-link{ text-decoration:none !important; color: inherit !important; display:block; }

.kpi{
  background: linear-gradient(180deg, rgba(255,255,255,.96), rgba(255,255,255,.86));
  border: 1px solid rgba(15,23,42,.10);
  border-radius: 18px;
  box-shadow: var(--shadow);
  padding: 16px 18px;
  position: relative;
  overflow:hidden;
  transition: transform .22s ease, box-shadow .22s ease, filter .22s ease;
  cursor: pointer;
}

.kpi::before{
  content:"";
  position:absolute; left:0; top:0; bottom:0; width: 5px;
  background: #cbd5e1;
}
.kpi.kpi-ibs::before{ background: var(--blue); }
.kpi.kpi-cbs::before{ background: var(--green); }
.kpi.kpi-cred::before{ background: var(--amber); }
.kpi.kpi-total::before{ background: var(--purple); }

.kpi::after{
  content:"";
  position:absolute;
  width: 240px; height: 240px;
  right:-90px; top:-110px;
  border-radius: 999px;
  opacity: .55;
  background: radial-gradient(circle at 30% 30%, rgba(37,99,235,.20), transparent 60%);
}
.kpi.kpi-cbs::after{ background: radial-gradient(circle at 30% 30%, rgba(22,163,74,.22), transparent 60%); }
.kpi.kpi-cred::after{ background: radial-gradient(circle at 30% 30%, rgba(245,158,11,.26), transparent 60%); }
.kpi.kpi-total::after{ background: radial-gradient(circle at 30% 30%, rgba(124,58,237,.22), transparent 60%); }

.kpi:hover{ transform: translateY(-6px); box-shadow: var(--shadow2); }
.kpi:active{ transform: translateY(-2px); box-shadow: var(--shadow); }

.kpi.is-active{
  outline: 3px solid rgba(15,23,42,.10);
  box-shadow: var(--shadow2);
  transform: translateY(-4px);
}

.kpi-head{ display:flex; align-items:flex-start; justify-content:space-between; gap: 12px; margin-bottom: 8px; position: relative; z-index: 1; }
.kpi-icon{
  width: 40px; height: 40px; border-radius: 14px;
  border: 1px solid rgba(15,23,42,.08);
  display:flex; align-items:center; justify-content:center;
  background: rgba(255,255,255,.72);
  box-shadow: 0 10px 25px rgba(2,6,23,.08);
}
.kpi-icon svg{ width: 18px; height: 18px; opacity:.95; }

.kpi .label{ color: var(--muted); font-size: .90rem; font-weight: 700; }
.kpi .value{ color: var(--ink); font-size: 1.75rem; font-weight: 900; letter-spacing: -0.02em; position: relative; z-index: 1; }
.kpi .sub{ color: var(--muted); font-size: .86rem; margin-top: 4px; position: relative; z-index: 1; }

/* Panels */
.panel-title{ display:flex; align-items:flex-start; gap: 10px; margin-bottom: 8px; }
.panel-title h3{ margin:0; font-size: 1.05rem; color: var(--ink) !important; }
.panel-title .hint{ color: var(--muted); font-size: 0.86rem; margin-top: 2px; }

.icon{
  width: 34px; height: 34px; border-radius: 12px;
  border: 1px solid rgba(15,23,42,.08);
  display:flex; align-items:center; justify-content:center;
  background: rgba(255,255,255,.78);
  box-shadow: 0 10px 25px rgba(2,6,23,.08);
}
.icon svg{ width: 18px; height: 18px; opacity:.95; }

.bar-track{ height: 10px; background: rgba(15,23,42,.06); border-radius: 999px; overflow:hidden; border: 1px solid rgba(15,23,42,.07);}
.bar-fill{ height:100%; border-radius: 999px; }
.bar-fill.ibs{ background: var(--blue); }
.bar-fill.cbs{ background: var(--green); }
.bar-fill.cred{ background: var(--amber); }

.bar-label{ display:flex; justify-content:space-between; align-items:center; font-size: 0.92rem; color: var(--muted); margin-bottom: 6px; }
.bar-foot{ display:flex; justify-content:space-between; align-items:center; margin-top: 10px; padding-top: 10px; border-top:1px solid rgba(15,23,42,.10); }
.badge-money{ font-weight: 900; }

/* Buttons */
.stButton>button, .stDownloadButton>button{
  background: linear-gradient(135deg, #111827, #0f172a) !important;
  color: #fff !important;
  border: 1px solid rgba(255,255,255,.10) !important;
  border-radius: 14px !important;
  padding: 10px 14px !important;
  font-weight: 900 !important;
  box-shadow: 0 14px 35px rgba(2,6,23,.20) !important;
  transition: transform .2s ease, box-shadow .2s ease, filter .2s ease !important;
}
.stButton>button:hover, .stDownloadButton>button:hover{
  transform: translateY(-2px) !important;
  box-shadow: 0 22px 55px rgba(2,6,23,.26) !important;
  filter: brightness(1.03) !important;
}
.stButton>button:active, .stDownloadButton>button:active{ transform: translateY(0px) !important; }

/* Inputs */
.stTextInput input, .stDateInput input{
  border-radius: 14px !important;
  border: 1px solid rgba(15,23,42,.12) !important;
  box-shadow: 0 10px 25px rgba(2,6,23,.06) !important;
}
.stSelectbox div[data-baseweb="select"] > div{
  border-radius: 14px !important;
  border: 1px solid rgba(15,23,42,.12) !important;
  box-shadow: 0 10px 25px rgba(2,6,23,.06) !important;
}

/* DataFrame */
.stDataFrame{
  border-radius: 16px !important;
  overflow:hidden !important;
  border: 1px solid rgba(15,23,42,.10) !important;
  box-shadow: 0 18px 45px rgba(2,6,23,.10) !important;
}

/* Uploader custom card */
.uploader-box{
  background: rgba(255,255,255,.06);
  border: 1px solid rgba(255,255,255,.10);
  border-radius: 18px;
  padding: 16px;
  box-shadow: 0 18px 40px rgba(0,0,0,.25);
}

/* === FIX: remove decorative giant icons === */
.kpi::after{ display: none !important; }

/* Tip (Dica importante) – premium + icon sized correctly */
.tip{
  display:flex;
  gap: 12px;
  align-items:flex-start;
  padding: 14px 16px;
  border-radius: 16px;
  background: #fff7ed;
  border: 1px solid rgba(180,83,9,.18);
  box-shadow: 0 12px 35px rgba(2,6,23,.06);
}
.tip strong{ display:block; color:#b45309; font-weight:900; margin-bottom:2px; }
.tip span{ color:#92400e !important; font-size:.92rem; }
.tip-icon{
  width: 36px;
  height: 36px;
  border-radius: 14px;
  background: rgba(245,158,11,.18);
  border: 1px solid rgba(245,158,11,.22);
  display:flex;
  align-items:center;
  justify-content:center;
  flex: 0 0 auto;
}
.tip-icon svg { width: 18px; height: 18px; }

/* ===== FIX UPLOAD ZONA BRANCA (SIDEBAR) ===== */
section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"]{
  background: rgba(255,255,255,.06) !important;
  border: 1px dashed rgba(255,255,255,.22) !important;
  border-radius: 18px !important;
  padding: 14px !important;
}

section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] *{
  color: rgba(255,255,255,.90) !important;
}

section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] svg{
  fill: rgba(255,255,255,.90) !important;
  color: rgba(255,255,255,.90) !important;
  opacity: 1 !important;
}

section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button{
  background: rgba(255,255,255,.10) !important;
  border: 1px solid rgba(255,255,255,.18) !important;
  color: rgba(255,255,255,.92) !important;
  border-radius: 12px !important;
}

/* ===== TABELA PREMIUM (igual vídeo) ===== */
.table-wrap{
  background: rgba(255,255,255,.92);
  border: 1px solid rgba(15,23,42,.10);
  border-radius: 18px;
  box-shadow: 0 18px 45px rgba(2,6,23,.10);
  padding: 16px;
  margin-top: 10px;
}


/* ===== GLOW NO CONTORNO DOS KPIs (hover por cor) ===== */
.kpi{
  transition: transform .22s ease, box-shadow .22s ease, border-color .22s ease;
}

/* IBS - azul */
.kpi.kpi-ibs:hover{
  box-shadow:
    0 26px 70px rgba(2,6,23,.16),
    0 0 0 1px rgba(37,99,235,.25),
    0 0 22px rgba(37,99,235,.35),
    0 0 60px rgba(37,99,235,.18) !important;
}

/* CBS - verde */
.kpi.kpi-cbs:hover{
  box-shadow:
    0 26px 70px rgba(2,6,23,.16),
    0 0 0 1px rgba(22,163,74,.25),
    0 0 22px rgba(22,163,74,.35),
    0 0 60px rgba(22,163,74,.18) !important;
}

/* Créditos - laranja */
.kpi.kpi-cred:hover{
  box-shadow:
    0 26px 70px rgba(2,6,23,.16),
    0 0 0 1px rgba(245,158,11,.30),
    0 0 22px rgba(245,158,11,.40),
    0 0 60px rgba(245,158,11,.18) !important;
}

/* Total - roxo */
.kpi.kpi-total:hover{
  box-shadow:
    0 26px 70px rgba(2,6,23,.16),
    0 0 0 1px rgba(124,58,237,.28),
    0 0 22px rgba(124,58,237,.38),
    0 0 60px rgba(124,58,237,.18) !important;
}


/* ===== GLOW NOS PAINÉIS DE DÉBITOS x CRÉDITOS ===== */
.card{
  transition: box-shadow .25s ease, transform .25s ease, border-color .25s ease;
}

/* IBS painel (azul) */
.card.ibs-panel:hover{
  box-shadow:
    0 20px 55px rgba(2,6,23,.18),
    0 0 0 1px rgba(37,99,235,.22),
    0 0 26px rgba(37,99,235,.30),
    0 0 70px rgba(37,99,235,.16) !important;
}

/* CBS painel (verde) */
.card.cbs-panel:hover{
  box-shadow:
    0 20px 55px rgba(2,6,23,.18),
    0 0 0 1px rgba(22,163,74,.22),
    0 0 26px rgba(22,163,74,.30),
    0 0 70px rgba(22,163,74,.16) !important;
}


/* ===== GLOW NA SIDEBAR (neon suave) ===== */
section[data-testid="stSidebar"]{
  box-shadow:
    0 30px 80px rgba(2,6,23,.45),
    0 0 0 1px rgba(99,102,241,.20),
    0 0 28px rgba(99,102,241,.35),
    0 0 90px rgba(99,102,241,.18) !important;
  transition: box-shadow .3s ease;
}

/* Intensifica levemente ao passar o mouse */
section[data-testid="stSidebar"]:hover{
  box-shadow:
    0 35px 95px rgba(2,6,23,.55),
    0 0 0 1px rgba(99,102,241,.28),
    0 0 36px rgba(99,102,241,.45),
    0 0 120px rgba(99,102,241,.22) !important;
}



/* ===== LOADER (UIVERSE SVG) – 4 CORES (IGUAL AOS CARDS) ===== */
:root{
  --ibs:#2563eb;   /* azul IBS */
  --cbs:#16a34a;   /* verde CBS */
  --cred:#f59e0b;  /* laranja Créditos */
  --total:#7c3aed; /* roxo Total */
}

/* some o ícone verde padrão */
div[data-testid="stStatusWidget"] { display:none !important; }

/* overlay premium */
.spinner-overlay{
  position: fixed;
  inset: 0;
  z-index: 99999;
  display:flex;
  align-items:center;
  justify-content:center;
  background: rgba(15,23,42,.22);
  backdrop-filter: blur(6px);
}

.spinner-card{
  width: min(520px, calc(100vw - 40px));
  border-radius: 22px;
  padding: 18px 18px 16px;
  background: linear-gradient(180deg, rgba(255,255,255,.92), rgba(255,255,255,.78));
  border: 1px solid rgba(15,23,42,.10);
  box-shadow: 0 26px 70px rgba(2,6,23,.22);
  display:flex;
  align-items:center;
  gap: 14px;
}

.pl{ width: 64px; height: 64px; flex: 0 0 auto; }

.pl__ring{ animation: ringA var(--speed, 2s) linear infinite; }
.pl__ring--a{ stroke: var(--c1); }
.pl__ring--b{ animation-name: ringB; stroke: var(--c2); }
.pl__ring--c{ animation-name: ringC; stroke: var(--c1); }
.pl__ring--d{ animation-name: ringD; stroke: var(--c2); }

/* textos */
.spinner-texts{ display:flex; flex-direction:column; gap: 3px; min-width:0; }
.spinner-title{
  font-weight: 900;
  color: #0f172a;
  letter-spacing: -.02em;
  font-size: 1.02rem;
  line-height: 1.1;
  white-space: nowrap;
  overflow: hidden;
  text-overflow: ellipsis;
}
.spinner-sub{
  font-size: .88rem;
  color: #64748b;
  display:flex;
  align-items:center;
  gap: 10px;
}
.spinner-pill{
  display:inline-flex;
  align-items:center;
  gap: 8px;
  padding: 6px 10px;
  border-radius: 999px;
  border: 1px solid rgba(15,23,42,.10);
  background: rgba(15,23,42,.04);
  font-weight: 900;
  color: #334155;
}
.spinner-dot{
  width:8px; height:8px; border-radius: 999px;
  background: var(--c1);
  box-shadow: 0 0 14px color-mix(in srgb, var(--c1) 55%, transparent);
}

/* presets de cor */
.spinner-ibs  { --c1: var(--ibs);  --c2: color-mix(in srgb, var(--ibs) 60%, #38bdf8); }
.spinner-cbs  { --c1: var(--cbs);  --c2: color-mix(in srgb, var(--cbs) 60%, #4ade80); }
.spinner-cred { --c1: var(--cred); --c2: color-mix(in srgb, var(--cred) 60%, #fbbf24); }
.spinner-total{ --c1: var(--total);--c2: color-mix(in srgb, var(--total) 60%, #a855f7); }

/* ===== ANIMAÇÕES do Uiverse (NAWSOME) ===== */
@keyframes ringA{
  from,4%{stroke-dasharray:0 660;stroke-width:20;stroke-dashoffset:-330}
  12%{stroke-dasharray:60 600;stroke-width:30;stroke-dashoffset:-335}
  32%{stroke-dasharray:60 600;stroke-width:30;stroke-dashoffset:-595}
  40%,54%{stroke-dasharray:0 660;stroke-width:20;stroke-dashoffset:-660}
  62%{stroke-dasharray:60 600;stroke-width:30;stroke-dashoffset:-665}
  82%{stroke-dasharray:60 600;stroke-width:30;stroke-dashoffset:-925}
  90%,to{stroke-dasharray:0 660;stroke-width:20;stroke-dashoffset:-990}
}
@keyframes ringB{
  from,12%{stroke-dasharray:0 220;stroke-width:20;stroke-dashoffset:-110}
  20%{stroke-dasharray:20 200;stroke-width:30;stroke-dashoffset:-115}
  40%{stroke-dasharray:20 200;stroke-width:30;stroke-dashoffset:-195}
  48%,62%{stroke-dasharray:0 220;stroke-width:20;stroke-dashoffset:-220}
  70%{stroke-dasharray:20 200;stroke-width:30;stroke-dashoffset:-225}
  90%{stroke-dasharray:20 200;stroke-width:30;stroke-dashoffset:-305}
  98%,to{stroke-dasharray:0 220;stroke-width:20;stroke-dashoffset:-330}
}
@keyframes ringC{
  from{stroke-dasharray:0 440;stroke-width:20;stroke-dashoffset:0}
  8%{stroke-dasharray:40 400;stroke-width:30;stroke-dashoffset:-5}
  28%{stroke-dasharray:40 400;stroke-width:30;stroke-dashoffset:-175}
  36%,58%{stroke-dasharray:0 440;stroke-width:20;stroke-dashoffset:-220}
  66%{stroke-dasharray:40 400;stroke-width:30;stroke-dashoffset:-225}
  86%{stroke-dasharray:40 400;stroke-width:30;stroke-dashoffset:-395}
  94%,to{stroke-dasharray:0 440;stroke-width:20;stroke-dashoffset:-440}
}
@keyframes ringD{
  from,8%{stroke-dasharray:0 440;stroke-width:20;stroke-dashoffset:0}
  16%{stroke-dasharray:40 400;stroke-width:30;stroke-dashoffset:-5}
  36%{stroke-dasharray:40 400;stroke-width:30;stroke-dashoffset:-175}
  44%,50%{stroke-dasharray:0 440;stroke-width:20;stroke-dashoffset:-220}
  58%{stroke-dasharray:40 400;stroke-width:30;stroke-dashoffset:-225}
  78%{stroke-dasharray:40 400;stroke-width:30;stroke-dashoffset:-395}
  86%,to{stroke-dasharray:0 440;stroke-width:20;stroke-dashoffset:-440}
}


/* ===== UIVERSE UPLOADER (mantém seu tema/cores) ===== */

/* “container” do uploader */
.uiverse-uploader section[data-testid="stFileUploaderDropzone"]{
  height: 300px !important;
  border-radius: 14px !important;
  box-shadow: 4px 4px 30px rgba(0,0,0,.20) !important;
  padding: 12px !important;
  gap: 8px !important;
  background: rgba(37,99,235,.06) !important;
  border: 1px solid rgba(255,255,255,.10) !important;
  position: relative !important;
  display: flex !important;
  flex-direction: column !important;
  justify-content: space-between !important;
}

/* “header” (área tracejada) */
.uiverse-uploader section[data-testid="stFileUploaderDropzone"] > div{
  flex: 1 !important;
  width: 100% !important;
  border: 2px dashed rgba(59,130,246,.55) !important;
  border-radius: 12px !important;
  display: flex !important;
  align-items: center !important;
  justify-content: center !important;
  flex-direction: column !important;
  background: rgba(255,255,255,.04) !important;
  padding: 14px !important;
}

/* Esconde o ícone padrão do Streamlit */
.uiverse-uploader section[data-testid="stFileUploaderDropzone"] svg{
  display:none !important;
}

/* Ícone novo (cloud upload) */
.uiverse-uploader section[data-testid="stFileUploaderDropzone"] > div::before{
  content:"";
  width: 92px;
  height: 92px;
  display:block;
  margin-bottom: 10px;
  background-repeat:no-repeat;
  background-size:contain;
  filter: drop-shadow(0 10px 24px rgba(37,99,235,.25));
  background-image: url("data:image/svg+xml;utf8,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%232563eb' stroke-width='1.8' stroke-linecap='round' stroke-linejoin='round'><path d='M20 16.58A5 5 0 0 0 18 7h-1.26A8 8 0 1 0 4 15.25'/><polyline points='16 16 12 12 8 16'/><line x1='12' y1='12' x2='12' y2='21'/></svg>");
}

/* Texto do dropzone */
.uiverse-uploader section[data-testid="stFileUploaderDropzone"] *{
  text-align:center !important;
}

/* “footer” (barra inferior + botão) */
.uiverse-uploader section[data-testid="stFileUploaderDropzone"] button{
  width: 100% !important;
  height: 42px !important;
  margin-top: 10px !important;
  border-radius: 12px !important;
  background: rgba(37,99,235,.10) !important;
  border: 1px solid rgba(59,130,246,.28) !important;
  color: rgba(226,232,240,.92) !important;
  font-weight: 900 !important;
  box-shadow: 0 2px 30px rgba(0,0,0,.18) !important;
}

.uiverse-uploader section[data-testid="stFileUploaderDropzone"] button:hover{
  filter: brightness(1.06) !important;
  transform: translateY(-1px) !important;
}

/* Mantém seu estilo escuro da sidebar */
section[data-testid="stSidebar"] .uiverse-uploader section[data-testid="stFileUploaderDropzone"]{
  background: rgba(255,255,255,.06) !important;
}


/* ===== FILE UPLOADER DA SIDEBAR – UIVERSE STYLE ===== */
section[data-testid="stSidebar"] section[data-testid="stFileUploaderDropzone"] {
  height: 300px !important;
  border-radius: 14px !important;
  box-shadow: 4px 4px 30px rgba(0,0,0,.20) !important;
  padding: 12px !important;
  background: rgba(37,99,235,.06) !important;
  border: 1px solid rgba(255,255,255,.10) !important;
  display: flex !important;
  flex-direction: column !important;
  justify-content: space-between !important;
}
section[data-testid="stSidebar"] section[data-testid="stFileUploaderDropzone"] > div {
  flex: 1 !important;
  width: 100% !important;
  border: 2px dashed rgba(59,130,246,.55) !important;
  border-radius: 12px !important;
  display: flex !important;
  align-items: center !important;
  justify-content: center !important;
  flex-direction: column !important;
  background: rgba(255,255,255,.04) !important;
}
section[data-testid="stSidebar"] section[data-testid="stFileUploaderDropzone"] svg {
  display: none !important;
}
section[data-testid="stSidebar"] section[data-testid="stFileUploaderDropzone"] > div::before {
  content: "";
  width: 92px;
  height: 92px;
  margin-bottom: 10px;
  background-repeat: no-repeat;
  background-size: contain;
  filter: drop-shadow(0 10px 24px rgba(37,99,235,.25));
  background-image: url("data:image/svg+xml;utf8,\
<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%232563eb' stroke-width='1.8' stroke-linecap='round' stroke-linejoin='round'>\
<path d='M20 16.58A5 5 0 0 0 18 7h-1.26A8 8 0 1 0 4 15.25'/>\
<polyline points='16 16 12 12 8 16'/>\
<line x1='12' y1='12' x2='12' y2='21'/>\
</svg>");
}
section[data-testid="stSidebar"] section[data-testid="stFileUploaderDropzone"] * {
  text-align: center !important;
}
section[data-testid="stSidebar"] section[data-testid="stFileUploaderDropzone"] button {
  width: 100% !important;
  height: 42px !important;
  border-radius: 12px !important;
  background: rgba(37,99,235,.10) !important;
  border: 1px solid rgba(59,130,246,.28) !important;
  font-weight: 900 !important;
}


/* ===== UPLOADER SIDEBAR – INTERATIVO (UX PREMIUM) ===== */

/* Mais compacto */
section[data-testid="stSidebar"]
section[data-testid="stFileUploaderDropzone"]{
  height: 180px !important;
  padding: 10px !important;
  border-radius: 16px !important;
  transition: background .22s ease, border-color .22s ease, box-shadow .22s ease, transform .22s ease;
}

/* Área tracejada */
section[data-testid="stSidebar"]
section[data-testid="stFileUploaderDropzone"] > div{
  padding: 10px !important;
  border-radius: 14px !important;
}

/* Esconde ícone padrão */
section[data-testid="stSidebar"]
section[data-testid="stFileUploaderDropzone"] svg{
  display: none !important;
}

/* Ícone cloud (tamanho + transição) */
section[data-testid="stSidebar"]
section[data-testid="stFileUploaderDropzone"] > div::before{
  width: 44px !important;
  height: 44px !important;
  margin-bottom: 6px !important;
  transition: transform .22s ease, filter .22s ease, opacity .22s ease;
  transform: translateY(0) scale(1);
  opacity: .92;
  filter: drop-shadow(0 6px 14px rgba(37,99,235,.22));
}

/* Hover: “chama” o usuário */
section[data-testid="stSidebar"]
section[data-testid="stFileUploaderDropzone"]:hover{
  background: rgba(37,99,235,.10) !important;
  border-color: rgba(59,130,246,.55) !important;
  box-shadow: 0 18px 45px rgba(2,6,23,.18), 0 0 0 1px rgba(59,130,246,.20), 0 0 22px rgba(59,130,246,.20) !important;
  transform: translateY(-1px);
}

/* Ícone: pulse no hover */
@keyframes uploadPulse {
  0%   { transform: translateY(0) scale(1); }
  50%  { transform: translateY(-2px) scale(1.08); }
  100% { transform: translateY(0) scale(1); }
}

section[data-testid="stSidebar"]
section[data-testid="stFileUploaderDropzone"]:hover > div::before{
  animation: uploadPulse .9s ease-in-out infinite;
  opacity: 1;
  filter: drop-shadow(0 10px 22px rgba(37,99,235,.34));
}

/* Clique: feedback */
section[data-testid="stSidebar"]
section[data-testid="stFileUploaderDropzone"]:active > div::before{
  animation: none !important;
  transform: translateY(1px) scale(.96) !important;
  filter: drop-shadow(0 4px 10px rgba(37,99,235,.25)) !important;
}

/* Texto mais compacto */
section[data-testid="stSidebar"]
section[data-testid="stFileUploaderDropzone"] p,
section[data-testid="stSidebar"]
section[data-testid="stFileUploaderDropzone"] small,
section[data-testid="stSidebar"]
section[data-testid="stFileUploaderDropzone"] span{
  font-size: .82rem !important;
  line-height: 1.15 !important;
  margin: 2px 0 !important;
  text-align: center !important;
}

/* Botão compacto */
section[data-testid="stSidebar"]
section[data-testid="stFileUploaderDropzone"] button{
  height: 34px !important;
  padding: 6px 10px !important;
  font-size: .88rem !important;
  border-radius: 12px !important;
}

/* Quando já tem arquivo: troca ícone para check verde */
section[data-testid="stSidebar"]
section[data-testid="stFileUploaderDropzone"]:has([data-testid="stFileUploaderFile"]){
  border-color: rgba(34,197,94,.55) !important;
  box-shadow: 0 18px 45px rgba(2,6,23,.18), 0 0 0 1px rgba(34,197,94,.18), 0 0 22px rgba(34,197,94,.18) !important;
}

section[data-testid="stSidebar"]
section[data-testid="stFileUploaderDropzone"]:has([data-testid="stFileUploaderFile"]) > div::before{
  animation: none !important;
  background-image: url("data:image/svg+xml;utf8,\
<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%2322c55e' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'>\
<polyline points='20 6 9 17 4 12'/>\
</svg>") !important;
  opacity: 1 !important;
  filter: drop-shadow(0 10px 22px rgba(34,197,94,.40)) !important;
}


/* ===== DOC TABLE PREMIUM (igual print) ===== */
.doc-table-wrap{
  background: rgba(255,255,255,.92);
  border: 1px solid rgba(15,23,42,.10);
  border-radius: 18px;
  box-shadow: 0 18px 45px rgba(2,6,23,.10);
  padding: 14px;
  overflow: hidden;
}

.doc-table{
  width: 100%;
  border-collapse: collapse;
  font-size: 12px;
}

.doc-table thead th{
  font-size: 10.5px;
  text-transform: uppercase;
  letter-spacing: .08em;
  font-weight: 900;
  color: rgba(100,116,139,.92);
  background: rgba(248,250,252,.95);
  padding: 12px 12px;
  border-bottom: 1px solid rgba(15,23,42,.10);
  text-align: left;
}

.doc-table tbody td{
  font-size: 12px;
  color: rgba(15,23,42,.92);
  padding: 12px 12px;
  border-bottom: 1px solid rgba(15,23,42,.07);
  vertical-align: middle;
}

.doc-table tbody tr:hover td{
  background: rgba(37,99,235,.04);
}

.doc-table .col-item{
  font-weight: 900;
  color: #0f172a;
}

.doc-table .col-money{
  text-align: right;
  font-variant-numeric: tabular-nums;
}

.doc-table .col-vibs{
  text-align: right;
  color: #2563eb;
  font-weight: 900;
  font-variant-numeric: tabular-nums;
}

.doc-table .col-vcbs{
  text-align: right;
  color: #16a34a;
  font-weight: 900;
  font-variant-numeric: tabular-nums;
}

.doc-table .col-file{
  max-width: 260px;
  white-space: nowrap;
  overflow: hidden;
  text-overflow: ellipsis;
  color: rgba(71,85,105,.95);
}

.cclass-badge{
  display:inline-flex;
  align-items:center;
  justify-content:center;
  font-size: 10.5px;
  padding: 4px 8px;
  border-radius: 999px;
  font-weight: 900;
  color: rgba(30,64,175,.95);
  background: rgba(37,99,235,.10);
  border: 1px solid rgba(37,99,235,.18);
}

.doc-table-foot{
  margin-top: 10px;
  font-size: .80rem;
  color: rgba(100,116,139,.95);
}


/* ===== DOC TABLE: ALTURA FIXA + SCROLL (não estica a página) ===== */
.doc-table-wrap{
  max-height: 520px !important;
  overflow: auto !important;
}

/* mantém cabeçalho “grudado” ao rolar */
.doc-table thead th{
  position: sticky !important;
  top: 0 !important;
  z-index: 2 !important;
}

/* Item/Serviço menos “grosso” */
.doc-table .col-item{
  font-weight: 800 !important;
}

/* scrollbar elegante */
.doc-table-wrap::-webkit-scrollbar{ width: 10px; height: 10px; }
.doc-table-wrap::-webkit-scrollbar-thumb{
  background: rgba(15,23,42,.14);
  border-radius: 999px;
}
.doc-table-wrap::-webkit-scrollbar-track{
  background: rgba(15,23,42,.05);
  border-radius: 999px;
}


/* ===== AJUSTES FINOS (data sem quebrar + espaço botão + chip arquivo bonito) ===== */

/* Data e Número: não quebrar */
.doc-table .col-date,
.doc-table .col-num{
  white-space: nowrap !important;
  font-variant-numeric: tabular-nums !important;
}

/* Também garante nas células, caso classe não esteja aplicada */
.doc-table tbody td.col-date,
.doc-table tbody td.col-num{
  white-space: nowrap !important;
}

/* Espaço pro botão “Baixar CSV filtrado” não encostar na tabela */
.table-download-spacer{
  height: 14px;
}

/* Chip do arquivo carregado (sidebar) – versão bonita compacta */
section[data-testid="stSidebar"] [data-testid="stFileUploaderFile"]{
  background: linear-gradient(180deg, rgba(255,255,255,.10), rgba(255,255,255,.06)) !important;
  border: 1px solid rgba(255,255,255,.14) !important;
  border-radius: 12px !important;
  padding: 6px 10px !important;
  margin-top: 8px !important;
  box-shadow:
    0 12px 28px rgba(0,0,0,.22),
    0 0 0 1px rgba(37,99,235,.10) !important;
  backdrop-filter: blur(10px) !important;
}

section[data-testid="stSidebar"] [data-testid="stFileUploaderFile"] *{
  color: rgba(226,232,240,.92) !important;
  line-height: 1.15 !important;
}

section[data-testid="stSidebar"] [data-testid="stFileUploaderFile"] svg{
  width: 16px !important;
  height: 16px !important;
  color: rgba(59,130,246,.95) !important;
  fill: rgba(59,130,246,.95) !important;
  filter: drop-shadow(0 8px 18px rgba(59,130,246,.18));
}

section[data-testid="stSidebar"] [data-testid="stFileUploaderFile"] button{
  border-radius: 8px !important;
  border: 1px solid rgba(255,255,255,.16) !important;
  background: rgba(255,255,255,.08) !important;
  padding: 3px 6px !important;
}

section[data-testid="stSidebar"] [data-testid="stFileUploaderFile"] button:hover{
  filter: brightness(1.12) !important;
  background: rgba(255,255,255,.12) !important;
}


/* ===== HARDEN UI (ocultar tudo que for chrome do Streamlit) ===== */
header[data-testid="stHeader"]{display:none !important;}
div[data-testid="stToolbar"]{display:none !important;}
div[data-testid="stDecoration"]{display:none !important;}
#MainMenu{display:none !important;}
footer{display:none !important;}
div[data-testid="stDeployButton"]{display:none !important;}
button[title="View fullscreen"], button[title="Exit fullscreen"]{display:none !important;}
/* alguns builds usam esses wrappers */
div[class*="stAppToolbar"], div[class*="stToolbar"]{display:none !important;}

/* remove espaço superior deixado pelo header */
.stApp .block-container{padding-top: 1rem !important;}

/* Streamlit Cloud/Share/Manage: remove qualquer widget FIXO no canto inferior direito */
div[style*="position: fixed"][style*="bottom"][style*="right"]{display:none !important;}
div[style*="position:fixed"][style*="bottom"][style*="right"]{display:none !important;}

/* Remover qualquer overlay de “status”/toast do Streamlit */
div[data-testid="stToast"]{display:none !important;}


/* ===== MICRO-INTERAÇÕES (premium | gira e para) ===== */

/* Base (efeito mola) */
.kpi .kpi-icon, .card .icon{
  transition: filter 0.35s ease;
  will-change: filter;
}
.kpi .kpi-icon svg, .card .icon svg{
  display: block;
  transform-origin: 50% 50%;
  transition: transform 0.75s cubic-bezier(.22,1.28,.42,1);
  will-change: transform;
}

/* Hover: gira UMA vez e para (bem visível) */
.kpi:hover .kpi-icon svg{
  transform: rotate(180deg) scale(1.08);
}
.card:hover .icon svg{
  transform: rotate(160deg) scale(1.06);
}

/* Glow suave */
.kpi:hover .kpi-icon{
  filter: drop-shadow(0 0 14px rgba(99,102,241,.35));
}
.card:hover .icon{
  filter: drop-shadow(0 0 12px rgba(99,102,241,.25));
}

/* Micro movimento na barra */
.card .bar-fill{
  transform-origin: left center;
  transition: transform .35s ease;
}
.card:hover .bar-fill{ transform: scaleX(1.03); }

/* ===== Upload: esconder Browse + card inteiro clicável ===== */
section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button,
section[data-testid="stSidebar"] [data-testid="stFileUploader"] button{
  display: none !important;
}
section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"]{
  cursor: pointer !important;
}

</style>
"""

st.markdown(CSS, unsafe_allow_html=True)


# -----------------------------
# HTML helper (avoids Markdown code-block due to indentation)
# -----------------------------
def _html_block(s: str):
    # Streamlit markdown treats leading 4 spaces as code block.
    # Dedent and strip to guarantee HTML renders.
    st.markdown(dedent(s).strip(), unsafe_allow_html=True)


def _html_clean(s: str) -> str:
    """Normaliza HTML para evitar que o Markdown do Streamlit transforme em bloco de código."""
    raw = dedent(s)
    lines = [ln.lstrip() for ln in raw.splitlines() if ln.strip()]
    return "\n".join(lines)


# -----------------------------
# Spinner overlay (4 cores)
# -----------------------------
spinner_placeholder = st.empty()

def spinner_html(tipo: str, titulo: str, subtitulo: str, speed: str = "2s") -> str:
    # Remove *qualquer* indentação para evitar o Markdown transformar em bloco de código
    raw = dedent(f"""<div class="spinner-overlay">
<div class="spinner-card spinner-{tipo}" style="--speed:{speed}">
<svg class="pl" viewBox="0 0 240 240" aria-hidden="true">
<circle class="pl__ring pl__ring--a" cx="120" cy="120" r="105" fill="none" stroke-width="20"/>
<circle class="pl__ring pl__ring--b" cx="120" cy="120" r="35"  fill="none" stroke-width="20"/>
<circle class="pl__ring pl__ring--c" cx="120" cy="120" r="70"  fill="none" stroke-width="20"/>
<circle class="pl__ring pl__ring--d" cx="120" cy="120" r="105" fill="none" stroke-width="20"/>
</svg>
<div class="spinner-texts">
<div class="spinner-title">{titulo}</div>
<div class="spinner-sub">
<span class="spinner-pill"><span class="spinner-dot"></span>{subtitulo}</span>
</div>
</div>
</div>
</div>""")
    return "\n".join(line.lstrip() for line in raw.splitlines() if line.strip())

def show_spinner(tipo: str, titulo: str, subtitulo: str, speed: str = "2s") -> None:
    spinner_placeholder.markdown(spinner_html(tipo, titulo, subtitulo, speed), unsafe_allow_html=True)

def hide_spinner() -> None:
    spinner_placeholder.empty()

# Spinner overlay (neon) – usado durante upload/processamento
spinner_placeholder = st.empty()
SPINNER_HTML = dedent("""
<div class="spinner-overlay">
  <div class="spinner-wrapper">
    <div class="spinner"></div>
    <div class="spinner1"></div>
  </div>
</div>
""")


# -----------------------------
# XML helpers
# -----------------------------
def _local(tag: str) -> str:
    # "{ns}Tag" -> "Tag"
    return tag.split("}", 1)[-1] if "}" in tag else tag

def _find_text(elem: ET.Element, path: str) -> str | None:
    x = elem.find(path)
    if x is None or x.text is None:
        return None
    return x.text.strip()

def _parse_date(root: ET.Element) -> date | None:
    """
    Tenta pegar data de emissão:
      - NFe/infNFe/ide/dhEmi (ISO datetime) ou dEmi (YYYY-MM-DD)
    """
    for p in [
        ".//{*}infNFe/{*}ide/{*}dhEmi",
        ".//{*}infNFe/{*}ide/{*}dEmi",
        ".//{*}ide/{*}dhEmi",
        ".//{*}ide/{*}dEmi",
    ]:
        t = _find_text(root, p)
        if not t:
            continue
        try:
            # dhEmi pode ser "2026-01-08T10:22:33-03:00"
            if "T" in t:
                # remove timezone para parse mais simples
                base = t.split("T")[0]
                return datetime.fromisoformat(base).date() if len(base) > 10 else datetime.fromisoformat(t[:19]).date()
            return datetime.fromisoformat(t).date()
        except Exception:
            try:
                return datetime.strptime(t[:10], "%Y-%m-%d").date()
            except Exception:
                pass
    return None

def _parse_nnf(root: ET.Element) -> str | None:
    # Número da NF: ide/nNF
    for p in [".//{*}infNFe/{*}ide/{*}nNF", ".//{*}ide/{*}nNF"]:
        t = _find_text(root, p)
        if t:
            return t
    return None


def _extract_nfe_key(xml_bytes: bytes) -> str:
    """Tenta extrair a chave (44 dígitos) da NFe/NFCe.
    - Prioriza Id do infNFe (ex.: Id="NFe3519...")
    - Fallback para tags chNFe comuns em protNFe/infProt ou eventos.
    Retorna "" se não encontrar.
    """
    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return ""

    # 1) infNFe @Id (mais comum)
    inf = root.find(".//{*}infNFe")
    if inf is not None:
        idv = inf.attrib.get("Id") or inf.attrib.get("id") or ""
        digits = "".join(ch for ch in idv if ch.isdigit())
        if len(digits) >= 44:
            return digits[-44:]

    # 2) chNFe em protocolos
    ch = (
        _find_text(root, ".//{*}protNFe/{*}infProt/{*}chNFe")
        or _find_text(root, ".//{*}infProt/{*}chNFe")
        or _find_text(root, ".//{*}chNFe")
        or ""
    )
    ch_digits = "".join(chh for chh in ch if chh.isdigit())
    if len(ch_digits) >= 44:
        return ch_digits[-44:]
    return ""


def _xml_signature(xml_bytes: bytes) -> str:
    """Assinatura estável para deduplicação:
    - Se achar chave, usa chave (melhor)
    - Senão, usa hash do conteúdo (sha1)
    """
    chave = _extract_nfe_key(xml_bytes)
    if chave:
        return f"ch:{chave}"
    return "sha1:" + hashlib.sha1(xml_bytes).hexdigest()

def _parse_items_from_xml(xml_bytes: bytes, filename: str) -> list[dict]:
    """
    Extrai itens (det) e IBS/CBS:
      - Item/Serviço: det/prod/xProd
      - cClassTrib: imposto/IBSCBS/cClassTrib
      - Base (vBC): imposto/IBSCBS/vBC
      - vIBS / vCBS: imposto/IBSCBS/vIBS, vCBS (se existirem)
    """
    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return []

    emissao = _parse_date(root)
    nnf = _parse_nnf(root)

    rows: list[dict] = []
    dets = root.findall(".//{*}infNFe/{*}det") or root.findall(".//{*}det")
    for det in dets:
        xprod = _find_text(det, ".//{*}prod/{*}xProd") or ""
        # Componentes do item (para validação por subtração)
        vprod = _find_text(det, ".//{*}prod/{*}vProd")
        vdesc = _find_text(det, ".//{*}prod/{*}vDesc")

        # Tributos por ITEM (quando existirem)
        vicms_item = _find_text(det, ".//{*}imposto/{*}ICMS//{*}vICMS")
        vpis_item = _find_text(det, ".//{*}imposto/{*}PIS//{*}vPIS")
        vcof_item = _find_text(det, ".//{*}imposto/{*}COFINS//{*}vCOFINS")

        ibscbs = det.find(".//{*}imposto/{*}IBSCBS")
        if ibscbs is None:
            # alguns XML podem não ter IBSCBS -> ignora item
            continue

        cclass = _find_text(ibscbs, ".//{*}cClassTrib") or ""
        vbc = _find_text(ibscbs, ".//{*}vBC")
        vibs = _find_text(ibscbs, ".//{*}vIBS")
        vcbs = _find_text(ibscbs, ".//{*}vCBS")

        def _to_float(x: str | None):
            try:
                if x in (None, ""):
                    return None
                # suporta vírgula decimal
                s = str(x).strip().replace(",", ".")
                return float(s)
            except Exception:
                return None

        def _to_float0(x: str | None) -> float:
            v = _to_float(x)
            return float(v) if v is not None else 0.0

        vbc_f = _to_float(vbc)
        vibs_f = _to_float(vibs)
        vcbs_f = _to_float(vcbs)

        # Componentes para validação por subtração (sempre em float)
        vprod_f = _to_float0(vprod)
        vdesc_f = _to_float0(vdesc)
        vicms_item_f = _to_float0(vicms_item)
        vpis_item_f = _to_float0(vpis_item)
        vcof_item_f = _to_float0(vcof_item)

        # Fonte do valor (base)
        fonte = "IBSCBS/vBC" if vbc_f is not None else ""

        rows.append(
            {
                "Data": emissao,
                "Numero": nnf,
                "Item/Serviço": xprod,
                "cClassTrib": cclass,
                "Valor da operação": vbc_f,
                "vIBS": vibs_f,
                "vCBS": vcbs_f,
                "vProd": vprod_f,
                "vDesc": vdesc_f,
                "vICMS_item": vicms_item_f,
                "vPIS_item": vpis_item_f,
                "vCOFINS_item": vcof_item_f,
                "arquivo": filename,
                "Fonte do valor": fonte,
            }
        )

    return rows



def _parse_tax_totals_from_xml(xml_bytes: bytes) -> dict:
    """Extrai totais do XML (por NOTA) via ICMSTot:
    - vICMS (ICMS próprio)
    - vPIS
    - vCOFINS
    """
    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return {"vICMS": 0.0, "vPIS": 0.0, "vCOFINS": 0.0}

    def _to_float(x: str | None) -> float:
        try:
            return float(x) if x not in (None, "") else 0.0
        except Exception:
            return 0.0

    vICMS = _find_text(root, ".//{*}ICMSTot/{*}vICMS")
    vPIS = _find_text(root, ".//{*}ICMSTot/{*}vPIS")
    vCOF = _find_text(root, ".//{*}ICMSTot/{*}vCOFINS")

    return {"vICMS": _to_float(vICMS), "vPIS": _to_float(vPIS), "vCOFINS": _to_float(vCOF)}


# ============================
# Validação Premium IBS/CBS
# Regra: Base Calc = vProd − vDesc − vICMS_item − vPIS_item − vCOFINS_item
# Zero tolerância: precisa bater exatamente (0,00).
# ============================

TOLERANCIA_BASE_IBSCBS = 0.0  # ZERO TOLERÂNCIA

def _br_money(v: float) -> str:
    try:
        s = f"{float(v):,.2f}"
        return s.replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "0,00"

def _safe_num(x) -> float:
    try:
        if x in (None, ""):
            return 0.0
        if isinstance(x, str):
            x = x.strip().replace(".", "").replace(",", ".")
        return float(x)
    except Exception:
        return 0.0

def aplicar_validacao_base_ibscbs(df_itens: pd.DataFrame) -> pd.DataFrame:
    """Adiciona colunas de validação IBS/CBS (por item)."""
    df = df_itens.copy()

    # Base do XML já vem em 'Valor da operação' (IBSCBS/vBC) no seu app
    if "Valor da operação" in df.columns:
        base_xml = df["Valor da operação"].fillna(0).apply(_safe_num)
    else:
        base_xml = pd.Series([0.0]*len(df), index=df.index)

    vProd = df.get("vProd", 0)
    vDesc = df.get("vDesc", 0)
    vICMS = df.get("vICMS_item", 0)
    vPIS = df.get("vPIS_item", 0)
    vCOF = df.get("vCOFINS_item", 0)

    vProd = pd.Series(vProd).fillna(0).apply(_safe_num)
    vDesc = pd.Series(vDesc).fillna(0).apply(_safe_num)
    vICMS = pd.Series(vICMS).fillna(0).apply(_safe_num)
    vPIS  = pd.Series(vPIS).fillna(0).apply(_safe_num)
    vCOF  = pd.Series(vCOF).fillna(0).apply(_safe_num)

    base_calc = (vProd - vDesc - vICMS - vPIS - vCOF).round(2)
    dif = (base_calc - base_xml).round(2)

    status = dif.apply(lambda d: "OK" if abs(d) <= TOLERANCIA_BASE_IBSCBS else "Divergente")

    df["Base IBS/CBS (XML)"] = base_xml.round(2)
    df["Base IBS/CBS (Calc)"] = base_calc
    df["Dif Base IBS/CBS"] = dif
    df["Status Base IBS/CBS"] = status

    # Diagnóstico curto (premium)
    def _diag(row):
        if row["Status Base IBS/CBS"] == "OK":
            return "✓ Base bateu exatamente (0,00)"
        # Se calc zerou mas XML > 0: normalmente faltam tributos por item (ou vProd não veio)
        if row["Base IBS/CBS (Calc)"] == 0 and row["Base IBS/CBS (XML)"] > 0:
            return "Componentes do item vieram 0,00 (ver vProd/vDesc/tributos por item)"
        return "Base do XML não bate com a decomposição do item (subtração)"

    df["Diagnóstico Base IBS/CBS"] = df.apply(_diag, axis=1)

    return df


def render_painel_validacao_premium(df_validado: pd.DataFrame, *, key_prefix: str = "ibscbs"):
    """Retângulo premium com resumo + cálculo detalhado.

    ✅ Fix:
    - Dropdown pode mostrar só divergentes
    - Painel de detalhe renderiza via components.html (não vira texto/código)
    - Botão para exportar apenas divergentes
    - Card fica vermelho quando item selecionado está divergente
    """
    if df_validado is None or len(df_validado) == 0:
        return

    # CSS premium (injetado uma vez)
    _html_block("""
<style>
/* ===== IBS/CBS Validation Panel (Neon Premium) ===== */
.ibscbs-panel{
  position:relative;
  background: linear-gradient(180deg, rgba(255,255,255,.96) 0%, rgba(255,255,255,.86) 100%);
  border: 1px solid rgba(148,163,184,.35);
  border-radius: 20px;
  padding: 18px;
  box-shadow: 0 18px 56px rgba(15,23,42,.10);
  backdrop-filter: blur(10px);
  overflow:hidden;
  transition: transform .22s ease, box-shadow .22s ease, border-color .22s ease;
}

/* animated neon border */
.ibscbs-panel::before{
  content:"";
  position:absolute; inset:-2px;
  border-radius: 22px;
  padding:2px;
  background: linear-gradient(90deg,
    rgba(37,99,235,.55),
    rgba(22,163,74,.55),
    rgba(245,158,11,.55),
    rgba(124,58,237,.55),
    rgba(37,99,235,.55)
  );
  background-size: 300% 300%;
  animation: ibscbsGlow 6s ease-in-out infinite;
  -webkit-mask:
    linear-gradient(#000 0 0) content-box,
    linear-gradient(#000 0 0);
  -webkit-mask-composite: xor;
  mask-composite: exclude;
  opacity:.55;
  pointer-events:none;
}
@keyframes ibscbsGlow{
  0%{background-position:0% 50%}
  50%{background-position:100% 50%}
  100%{background-position:0% 50%}
}

.ibscbs-panel:hover{
  transform: translateY(-2px);
  box-shadow:
    0 24px 70px rgba(15,23,42,.14),
    0 0 0 1px rgba(99,102,241,.14),
    0 0 34px rgba(99,102,241,.18);
}

.ibscbs-panel.divergente{
  border-color: rgba(239,68,68,.30);
  box-shadow:
    0 18px 56px rgba(15,23,42,.10),
    0 0 0 1px rgba(239,68,68,.18),
    0 0 30px rgba(239,68,68,.14);
}
.ibscbs-panel.divergente::before{ opacity:.70; }

/* Header */
.ibscbs-header{display:flex;align-items:flex-start;justify-content:space-between;gap:12px;margin-bottom:12px}
.ibscbs-title{display:flex;align-items:flex-start;gap:10px}
.ibscbs-title h3{margin:0;font-size:16px;font-weight:900;color:#0f172a;letter-spacing:-.2px}
.ibscbs-title p{margin:3px 0 0 0;font-size:12px;color:#64748b;max-width:860px}

/* Chip */
.ibscbs-chip{
  display:inline-flex;align-items:center;gap:8px;
  padding:8px 12px;border-radius:999px;font-size:12px;font-weight:900;
  border:1px solid transparent;white-space:nowrap;
  box-shadow: 0 10px 26px rgba(2,6,23,.08);
}
.ibscbs-chip.ok{color:#15803d;background:rgba(34,197,94,.14);border-color:rgba(34,197,94,.24)}
.ibscbs-chip.bad{color:#b91c1c;background:rgba(239,68,68,.14);border-color:rgba(239,68,68,.24)}

/* Metrics */
.ibscbs-metrics{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px;margin-top:10px;margin-bottom:14px}
.ibscbs-metric{
  background: rgba(248,250,252,.92);
  border:1px solid rgba(226,232,240,.95);
  border-radius:16px;
  padding:12px;
  transition: transform .22s ease, box-shadow .22s ease, border-color .22s ease;
}
.ibscbs-metric:hover{
  transform: translateY(-2px);
  box-shadow: 0 18px 45px rgba(15,23,42,.10), 0 0 22px rgba(99,102,241,.12);
  border-color: rgba(99,102,241,.22);
}
.ibscbs-metric .k{font-size:12px;color:#64748b;margin:0}
.ibscbs-metric .v{font-size:18px;font-weight:950;color:#0f172a;margin:6px 0 0 0}
.ibscbs-metric .s{font-size:11px;color:#94a3b8;margin:6px 0 0 0}
.ibscbs-divider{height:1px;background:rgba(226,232,240,.95);margin:14px 0}

/* Calc layout */
.ibscbs-calc{display:grid;grid-template-columns:1.15fr .85fr;gap:12px}

.calc-left, .calc-right{
  background: rgba(248,250,252,.92);
  border: 1px solid rgba(226,232,240,.95);
  border-radius: 16px;
  padding: 14px;
  box-shadow: 0 16px 40px rgba(2,6,23,.06);
  position:relative;
  overflow:hidden;
}

.calc-left::after, .calc-right::after{
  content:"";
  position:absolute;
  width: 260px; height: 260px;
  right:-120px; top:-140px;
  border-radius: 999px;
  background: radial-gradient(circle at 30% 30%, rgba(99,102,241,.18), transparent 60%);
  pointer-events:none;
}

.calc-head{
  display:flex;
  align-items:center;
  justify-content:space-between;
  gap:10px;
  margin-bottom:10px;
}
.calc-head .t{
  font-size:12px;
  font-weight:950;
  letter-spacing:.06em;
  text-transform:uppercase;
  color:#334155;
}
.calc-badge{
  display:inline-flex;align-items:center;gap:8px;
  padding:6px 10px;border-radius:999px;
  font-size:11px;font-weight:950;
  border:1px solid rgba(148,163,184,.25);
  background: rgba(15,23,42,.04);
  color:#334155;
}
.calc-badge .dot{
  width:8px;height:8px;border-radius:999px;background:rgba(99,102,241,.95);
  box-shadow: 0 0 14px rgba(99,102,241,.30);
}

.calc-lines{
  display:flex;
  flex-direction:column;
  gap:8px;
  margin-top:8px;
}

.calc-line{
  display:flex;
  align-items:center;
  justify-content:space-between;
  gap:10px;
  padding:10px 10px;
  border-radius:14px;
  border:1px solid rgba(226,232,240,.95);
  background: rgba(255,255,255,.72);
  transition: transform .18s ease, box-shadow .18s ease, border-color .18s ease;
}
.calc-line:hover{
  transform: translateY(-1px);
  border-color: rgba(99,102,241,.25);
  box-shadow: 0 14px 30px rgba(2,6,23,.08), 0 0 18px rgba(99,102,241,.12);
}
.calc-line .name{
  font-weight:950;
  color:#0f172a;
  display:flex;
  align-items:center;
  gap:8px;
}
.calc-line .name i{
  display:inline-flex;
  width: 22px; height: 22px;
  align-items:center; justify-content:center;
  border-radius: 9px;
  border:1px solid rgba(148,163,184,.25);
  background: rgba(15,23,42,.04);
  font-style: normal;
}
.calc-line .val{
  font-variant-numeric: tabular-nums;
  font-weight:950;
  color:#0f172a;
}
.calc-line.minus .name i{ border-color: rgba(245,158,11,.25); background: rgba(245,158,11,.10); }
.calc-line.icms .name i{ border-color: rgba(37,99,235,.25); background: rgba(37,99,235,.10); }
.calc-line.pis  .name i{ border-color: rgba(124,58,237,.25); background: rgba(124,58,237,.10); }
.calc-line.cof  .name i{ border-color: rgba(22,163,74,.25); background: rgba(22,163,74,.10); }

.calc-eq{
  margin-top:10px;
  padding-top:10px;
  border-top: 1px dashed rgba(148,163,184,.35);
  display:flex;
  align-items:center;
  justify-content:space-between;
  gap:10px;
}
.calc-eq .eq{
  font-family: ui-monospace,SFMono-Regular,Menlo,Monaco,Consolas,"Liberation Mono","Courier New",monospace;
  font-size:12px;
  color:#334155;
}
.calc-eq .res{
  font-weight: 950;
  color:#0f172a;
  font-variant-numeric: tabular-nums;
}

/* Right side rows */
.calc-right .row{
  display:flex;
  align-items:center;
  justify-content:space-between;
  gap:10px;
  margin:8px 0;
  padding:10px 10px;
  border-radius:14px;
  border:1px solid rgba(226,232,240,.95);
  background: rgba(255,255,255,.72);
  transition: transform .18s ease, box-shadow .18s ease, border-color .18s ease;
}
.calc-right .row:hover{
  transform: translateY(-1px);
  border-color: rgba(99,102,241,.22);
  box-shadow: 0 14px 30px rgba(2,6,23,.08), 0 0 18px rgba(99,102,241,.10);
}
.calc-right .row span{font-size:12px;color:#64748b}
.calc-right .row b{font-size:13px;color:#0f172a;font-variant-numeric: tabular-nums}

.calc-right .delta{
  margin-top:10px;
  padding-top:10px;
  border-top:1px dashed rgba(148,163,184,.35);
}

.calc-right .status-ok b{ color:#15803d; }
.calc-right .status-bad b{ color:#b91c1c; }

/* Footer */
.ibscbs-foot{margin-top:10px;font-size:11px;color:#94a3b8}

@media (max-width:900px){
  .ibscbs-metrics{grid-template-columns:repeat(2,minmax(0,1fr))}
  .ibscbs-calc{grid-template-columns:1fr}
}
</style>
""")

    total = len(df_validado)
    ok = int((df_validado["Status Base IBS/CBS"] == "OK").sum())
    div = total - ok

    soma_xml = float(df_validado["Base IBS/CBS (XML)"].sum())
    soma_calc = float(df_validado["Base IBS/CBS (Calc)"].sum())
    delta_total = round(soma_calc - soma_xml, 2)

    status_global_ok = (div == 0)
    chip = "ok" if status_global_ok else "bad"
    chip_txt = "✓ Validado (0,00)" if status_global_ok else f"⚠ Divergências ({div})"

    # Exportar só divergentes
    df_div = df_validado[df_validado["Status Base IBS/CBS"] != "OK"].copy()
    if not df_div.empty:
        csv_div = df_div.to_csv(index=False, sep=';', encoding='utf-8')
        st.download_button(
            "⬇️ Baixar somente divergentes (CSV)",
            data=csv_div,
            file_name="divergentes_ibscbs.csv",
            mime="text/csv",
            key=f"{key_prefix}_dl_div"
        )

    # Download do XML da nota selecionada (individual)
    try:
        sig_sel = str(row.get("xml_sig", "")).strip()
        store = st.session_state.get("xml_store", {})
        if sig_sel and sig_sel in store:
            meta = store[sig_sel]
            nnf = meta.get("Numero") or row.get("Numero") or ""
            chave = meta.get("chave") or ""
            fname = f"NFe_{nnf}.xml" if nnf else "nota.xml"
            if chave:
                fname = f"NFe_{nnf}_{chave[-6:]}.xml" if nnf else f"NFe_{chave[-6:]}.xml"
            st.download_button(
                "⬇️ Baixar XML desta nota",
                data=meta["bytes"],
                file_name=fname,
                mime="application/xml",
                key=f"{key_prefix}_dl_xml_{sig_sel}",
            )
    except Exception:
        pass

    # Dropdown: por padrão, só divergentes quando existir
    show_only_div = st.checkbox(
        "Mostrar somente as divergentes",
        value=(not df_div.empty),
        key=f"{key_prefix}_onlydiv",
        help="Filtra o seletor e mostra apenas itens com Status = Divergente."
    )

    df_tmp = df_validado.copy()
    if show_only_div:
        df_tmp = df_tmp[df_tmp["Status Base IBS/CBS"] != "OK"].copy()

    if df_tmp.empty:
        st.success("✅ Nenhuma divergência encontrada. (Tudo OK)")
        return

    df_tmp["_absdif"] = df_tmp["Dif Base IBS/CBS"].abs()
    df_tmp = df_tmp.sort_values("_absdif", ascending=False)

    label_col = "Item/Serviço" if "Item/Serviço" in df_tmp.columns else df_tmp.columns[0]
    options = df_tmp[label_col].fillna("").astype(str).tolist()

    pick = st.selectbox(
        "Detalhar cálculo (selecione um item)",
        options=options,
        index=0,
        key=f"{key_prefix}_pick",
        help="Mostra a decomposição do item: vProd − vDesc − ICMS_item − PIS_item − COFINS_item."
    )

    row = df_tmp[df_tmp[label_col].astype(str) == str(pick)].iloc[0]

    vProd = _safe_num(row.get("vProd"))
    vDesc = _safe_num(row.get("vDesc"))
    vICMS = _safe_num(row.get("vICMS_item"))
    vPIS  = _safe_num(row.get("vPIS_item"))
    vCOF  = _safe_num(row.get("vCOFINS_item"))

    base_xml = float(row["Base IBS/CBS (XML)"])
    base_calc = float(row["Base IBS/CBS (Calc)"])
    dif = float(row["Dif Base IBS/CBS"])

    status_item = "OK" if abs(dif) <= TOLERANCIA_BASE_IBSCBS else "Divergente"
    panel_class = "ibscbs-panel" + (" divergente" if status_item != "OK" else "")
    formula = (
        f"vProd ({_br_money(vProd)})  −  vDesc ({_br_money(vDesc)})  −  ICMS ({_br_money(vICMS)})  −  PIS ({_br_money(vPIS)})  −  COFINS ({_br_money(vCOF)})\n"
        f"= Base Calc ({_br_money(base_calc)})"
    )

    panel = f"""
<div class="{panel_class}">
  <div class="ibscbs-header">
    <div class="ibscbs-title">
      <div style="font-size:18px;">🧾</div>
      <div>
        <h3>Validação da Base IBS/CBS (ZERO tolerância)</h3>
        <p>Validação por subtração (item a item). A base calculada deve bater exatamente com a base do XML (IBSCBS/vBC). Qualquer centavo vira divergência.</p>
      </div>
    </div>
    <div class="ibscbs-chip {chip}">{chip_txt}</div>
  </div>

  <div class="ibscbs-metrics">
    <div class="ibscbs-metric"><p class="k">Itens</p><p class="v">{total}</p><p class="s">Total analisado</p></div>
    <div class="ibscbs-metric"><p class="k">Soma Base (XML)</p><p class="v">R$ {_br_money(soma_xml)}</p><p class="s">Total do XML</p></div>
    <div class="ibscbs-metric"><p class="k">Soma Base (Calc)</p><p class="v">R$ {_br_money(soma_calc)}</p><p class="s">Subtração por item</p></div>
    <div class="ibscbs-metric"><p class="k">Diferença</p><p class="v">R$ {_br_money(delta_total)}</p><p class="s">Calc − XML</p></div>
  </div>

  <div class="ibscbs-divider"></div>

  <div class="ibscbs-calc">
  <div class="calc-left">
    <div class="calc-head">
      <div class="t">Memória de cálculo</div>
      <div class="calc-badge"><span class="dot"></span>{status_item}</div>
    </div>

    <div class="calc-lines">
      <div class="calc-line">
        <div class="name"><i>+</i>vProd</div>
        <div class="val">R$ {_br_money(vProd)}</div>
      </div>

      <div class="calc-line minus">
        <div class="name"><i>−</i>vDesc</div>
        <div class="val">R$ {_br_money(vDesc)}</div>
      </div>

      <div class="calc-line icms">
        <div class="name"><i>−</i>ICMS</div>
        <div class="val">R$ {_br_money(vICMS)}</div>
      </div>

      <div class="calc-line pis">
        <div class="name"><i>−</i>PIS</div>
        <div class="val">R$ {_br_money(vPIS)}</div>
      </div>

      <div class="calc-line cof">
        <div class="name"><i>−</i>COFINS</div>
        <div class="val">R$ {_br_money(vCOF)}</div>
      </div>
    </div>

    <div class="calc-eq">
      <div class="eq">= Base Calc</div>
      <div class="res">R$ {_br_money(base_calc)}</div>
    </div>
  </div>

  <div class="calc-right">
    <div class="row"><span>Base XML</span><b>R$ {_br_money(base_xml)}</b></div>
    <div class="row"><span>Base Calc</span><b>R$ {_br_money(base_calc)}</b></div>
    <div class="row"><span>Diferença</span><b>R$ {_br_money(dif)}</b></div>

    <div class="delta">
      <div class="row {('status-ok' if status_item=='OK' else 'status-bad')}"><span>Status do item</span><b>{status_item}</b></div>
      <div class="row"><span>Nº da nota</span><b>{_h(str(row.get('Numero','') or ''))}</b></div>
      <div class="row"><span>Arquivo</span><b>{_h(row.get('arquivo',''))}</b></div>
    </div>
  </div>
</div>

<div class="ibscbs-foot">Regra rígida: diferença precisa ser <b>0,00</b>. Qualquer centavo vira divergência.</div>
</div>
"""

    # Render no corpo do Streamlit (aplicando CSS do app) sem virar texto.
    st.markdown(_html_clean(panel), unsafe_allow_html=True)


def _detect_cancel_event(xml_bytes: bytes) -> dict | None:
    """Detecta XML de evento de cancelamento (procEventoNFe / evento).
    Retorna dict com dados úteis ou None se não for cancelamento.
    """
    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return None

    # Procura tpEvento=110111 (Cancelamento)
    tp = _find_text(root, ".//{*}detEvento/{*}tpEvento") or _find_text(root, ".//{*}tpEvento")
    if tp != "110111":
        return None

    ch = _find_text(root, ".//{*}infEvento/{*}chNFe") or _find_text(root, ".//{*}chNFe") or ""
    dh = _find_text(root, ".//{*}infEvento/{*}dhEvento") or _find_text(root, ".//{*}dhEvento") or ""
    nprot = _find_text(root, ".//{*}infEvento/{*}nProt") or _find_text(root, ".//{*}nProt") or ""
    xjust = _find_text(root, ".//{*}detEvento/{*}xJust") or _find_text(root, ".//{*}xJust") or ""

    return {"chNFe": ch, "dhEvento": dh, "nProt": nprot, "xJust": xjust}



# -----------------------------
# Excel write helper
# -----------------------------
def _append_to_workbook(template_bytes: bytes, df: pd.DataFrame) -> bytes:
    """
    Abre o template e grava df na aba LANCAMENTOS, acrescentando linhas.

    ✅ O que este writer garante:
      - Encontra a linha correta de cabeçalhos mesmo que o layout mude (ex.: cabeçalho na linha 2).
      - Escreve nos campos de entrada (Data, Numero, Item/Serviço, etc.).
      - COPIA fórmulas/estilos da primeira linha-modelo de dados para todas as novas linhas,
        para que "Base", "Valor IBS/CBS", validações e cálculos voltem a aparecer no Excel.
    """
    from copy import copy
    bio = io.BytesIO(template_bytes)
    wb = load_workbook(bio)

    ws = wb["LANCAMENTOS"] if "LANCAMENTOS" in wb.sheetnames else wb.active

    # ------------------------------------------------------------
    # 1) Descobre em qual linha estão os cabeçalhos (layout pode mudar)
    # ------------------------------------------------------------
    expected = {"Data", "Numero", "Item/Serviço", "cClassTrib", "Valor da operação"}
    header_row = None

    # procura nos primeiros 25 rows (suficiente pro seu layout)
    for r in range(1, 26):
        values = []
        for c in range(1, 101):  # lê até 100 colunas (bem além do necessário)
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                values.append(v.strip())
        hit = len(expected.intersection(values))
        if hit >= 3:  # achou linha com a maioria dos cabeçalhos
            header_row = r
            break

    if header_row is None:
        # fallback antigo (assume linha 1)
        header_row = 1

    # mapeia "nome do cabeçalho" -> coluna
    headers: dict[str, int] = {}
    last_col = 0
    for col in range(1, 201):  # até 200 colunas
        v = ws.cell(row=header_row, column=col).value
        if isinstance(v, str) and v.strip():
            headers[v.strip()] = col
            last_col = max(last_col, col)

    # se ainda não achou nada (planilha muito custom), tenta usar as colunas usadas do sheet
    if last_col == 0:
        last_col = min(ws.max_column, 200)

    # ------------------------------------------------------------
    # 2) Define a "linha modelo" (a primeira linha de dados com fórmulas)
    #    No seu modelo: header_row=2, a linha 3 é seção, a 4 é a linha modelo.
    # ------------------------------------------------------------
    template_row = header_row + 2

    # ------------------------------------------------------------
    # 3) Descobre a próxima linha vazia olhando a coluna "Data"
    # ------------------------------------------------------------
    next_row = ws.max_row + 1
    if "Data" in headers:
        c = headers["Data"]
        r = ws.max_row
        while r >= (template_row) and ws.cell(row=r, column=c).value in (None, ""):
            r -= 1
        next_row = max(r + 1, template_row)

    # ------------------------------------------------------------
    # 4) Função para copiar estilo + fórmulas da linha modelo
    # ------------------------------------------------------------
    def _copy_row_style_and_formulas(src_row: int, dst_row: int):
        from copy import copy
        from openpyxl.formula.translate import Translator

        for col in range(1, last_col + 1):
            src = ws.cell(row=src_row, column=col)
            dst = ws.cell(row=dst_row, column=col)

            # estilos
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)

            # valor / fórmula
            if isinstance(src.value, str) and src.value.startswith("="):
                # traduz a referência da linha-modelo -> linha destino (ex.: G4 vira G7)
                try:
                    dst.value = Translator(src.value, origin=src.coordinate).translate_formula(dst.coordinate)
                except Exception:
                    dst.value = src.value
            else:
                dst.value = src.value


    # ------------------------------------------------------------
    # 5) Escreve as linhas: primeiro replica modelo, depois grava os valores de entrada
    # ------------------------------------------------------------
    fields = [
        "Data", "Numero", "Item/Serviço", "cClassTrib",
        "Valor da operação", "vIBS", "vCBS", "arquivo", "Fonte do valor"
    ]

    for _, row in df.iterrows():
        # replica a linha modelo (fórmulas + visual)
        _copy_row_style_and_formulas(template_row, next_row)

        # agora sobrescreve somente os campos de ENTRADA
        for f in fields:
            if f not in headers:
                continue
            col = headers[f]
            val = row.get(f, None)

            cell = ws.cell(row=next_row, column=col)

            # datas
            if f == "Data" and pd.notna(val) and isinstance(val, date):
                cell.value = val
                cell.number_format = "dd/mm/yyyy"
            else:
                if pd.isna(val):
                    val = None
                cell.value = val

        next_row += 1

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

# -----------------------------
# UI
# -----------------------------
# -----------------------------
# Header (modern - compact)
# -----------------------------
st.markdown("""
<style>
.header-container{
  font-family:'Inter',sans-serif;
  background:linear-gradient(135deg,rgba(255,255,255,.96),rgba(255,255,255,.88));
  border:1px solid rgba(15,23,42,.08);
  border-radius:16px;
  padding:14px 16px;
  margin:0 0 14px 0;
  box-shadow:0 12px 32px rgba(2,6,23,.10);
}
.header-top{display:flex;align-items:center;justify-content:space-between;gap:12px;}
.header-left{display:flex;align-items:center;gap:10px;flex-wrap:wrap;}
.version-badge{
  padding:4px 10px;border-radius:999px;
  font-size:11px;font-weight:800;
  background:rgba(59,130,246,.12);
  color:#2563eb;border:1px solid rgba(59,130,246,.25);
  white-space:nowrap;
}
.header-title{font-size:1.35rem;font-weight:900;margin:0;color:#0f172a;letter-spacing:-.02em;}
.header-title span{
  background:linear-gradient(135deg,#3b82f6,#10b981);
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;
}
.header-sub{font-size:.82rem;color:#64748b;margin-top:2px;}
.status-badge{
  display:inline-flex;align-items:center;gap:6px;
  padding:6px 10px;border-radius:999px;
  font-size:11px;font-weight:800;
  background:rgba(241,245,249,.9);
  border:1px solid #e2e8f0;color:#475569;
  white-space:nowrap;
}
.status-dot{width:7px;height:7px;border-radius:999px;background:#22c55e;}
.info-banner{
  margin-top:10px;
  padding:10px 12px;
  border-radius:12px;
  background:rgba(59,130,246,.08);
  border:1px solid rgba(59,130,246,.20);
  font-size:.82rem;color:#1e293b;
}
.info-banner b{color:#2563eb;font-weight:900;}
@media(max-width:820px){
  .header-top{flex-direction:column;align-items:flex-start;}
}
/* ===== Upload: esconder Browse + card inteiro clicável ===== */
section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button,
section[data-testid="stSidebar"] [data-testid="stFileUploader"] button{
  display: none !important;
}
section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"]{
  cursor: pointer !important;
}

</style>

<div class="header-container">
  <div class="header-top">
    <div>
      <div class="header-left">
        <span class="version-badge">⚡ v2.0</span>
        <h1 class="header-title">Extrator XML - <span>IBS/CBS</span></h1>
      </div>
      <div class="header-sub">🛡️ Visualização de dados fiscais da reforma tributária</div>
    </div>
    <div class="status-badge"><span class="status-dot"></span>Conectado</div>
  </div>

  <div class="info-banner">
    ⚡ Envie seus <b>XMLs</b> na lateral para inserir os dados na aba <b>LANÇAMENTOS</b> mantendo suas fórmulas.
  </div>
</div>
""", unsafe_allow_html=True)

# Sidebar: uploads
with st.sidebar:
    st.markdown(dedent("""
<div style="padding: 10px 6px 6px 6px;">
  <div style="display:flex; align-items:center; gap:10px; margin-bottom: 10px;">
<div style="width:42px;height:42px;border-radius:14px;background:rgba(108,124,255,.18);border:1px solid rgba(255,255,255,.12);display:flex;align-items:center;justify-content:center;">
<span style="font-weight:900;">✦</span>
</div>
<div>
<div style="font-weight:900; font-size: 1.05rem; line-height:1;">Extrator XML</div>
<div style="font-size:.82rem; color: rgba(226,232,240,.75); margin-top:2px;">IBS/CBS</div>
</div>
<div style="margin-left:auto; font-size:.75rem; font-weight:800; padding:4px 10px; border-radius:999px; background:rgba(255,255,255,.10); border:1px solid rgba(255,255,255,.10); color: rgba(226,232,240,.9);">v2.0</div>
  </div>

  <div class="sidebar-card">
<div class="sidebar-title">
<h4>EXCEL IBS/CBS</h4>
<span class="tag">FIXA</span>
</div>

<div class="uploader-box">
<div style="text-align:center; font-weight:800; margin-bottom: 6px;">Planilha interna</div>
<div style="text-align:center; color: rgba(226,232,240,.75); font-size:.82rem;">O app usa <b>planilha_modelo.xlsx</b></div>
</div>

<div style="height: 14px;"></div>

<div class="sidebar-title" style="margin-top: 2px;">

<h4>ARQUIVOS XML</h4>
<span class="tag" style="background:rgba(37,99,235,.18); border-color: rgba(37,99,235,.22);">OBRIGATÓRIO</span>
</div>

<div class="uploader-box">
"""), unsafe_allow_html=True)

    st.markdown('<div class="uiverse-uploader">', unsafe_allow_html=True)
    xml_files = st.file_uploader("", type=["xml","zip"], accept_multiple_files=True, label_visibility="collapsed")
    components.html(
        '''
    <script>
    (function(){
      function hook(){
        const dz = window.parent.document.querySelector('[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"]');
        if(!dz) return;
        dz.addEventListener('click', function(){
          const input = dz.querySelector('input[type="file"]');
          if(input) input.click();
        });
      }
      hook();
      setTimeout(hook, 800);
    })();
    </script>
        ''',
        height=0,
    )
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown(dedent("""
<div class="uploader-help">XML, ZIP • Múltiplos</div>
</div>

<div style="height: 14px;"></div>

<div style="padding: 12px; border-radius: 16px; background: rgba(255,255,255,.05); border: 1px solid rgba(255,255,255,.10);">
<div style="font-weight:900; margin-bottom: 4px;">Dica rápida</div>
<div style="font-size: .82rem; color: rgba(226,232,240,.78);">
        Envie XMLs para extrair automaticamente dados de IBS e CBS da reforma tributária.
</div>
</div>
  </div>
</div>
"""), unsafe_allow_html=True)

# Carrega planilha modelo FIXA (arquivo na pasta do projeto)
from pathlib import Path
TEMPLATE_PATH = Path(__file__).parent / "planilha_modelo.xlsx"

try:
    template_bytes = TEMPLATE_PATH.read_bytes()
except FileNotFoundError:
    template_bytes = None

# Aviso amigável caso o arquivo não exista (em produção ele deve estar junto do app)
if template_bytes is None:
    st.markdown(dedent("""
<div class="tip">
  <div class="tip-icon" aria-hidden="true">
    <svg viewBox="0 0 24 24" fill="none">
      <path d="M12 8v5" stroke="#b45309" stroke-width="1.8" stroke-linecap="round"/>
      <path d="M12 16h.01" stroke="#b45309" stroke-width="2.8" stroke-linecap="round"/>
      <path d="M10.3 3.7a2 2 0 0 1 3.4 0l8.4 14.7A2 2 0 0 1 20.4 21H3.6a2 2 0 0 1-1.7-3.0l8.4-14.3z"
            stroke="#b45309" stroke-width="1.6" fill="#fff7ed"/>
    </svg>
  </div>
  <div>
    <div class="tip-title">Planilha modelo não encontrada</div>
    <div class="tip-text">Coloque o arquivo <b>planilha_modelo.xlsx</b> na mesma pasta do <b>app.py</b>.</div>
  </div>
</div>
"""), unsafe_allow_html=True)

# Parse XMLs
rows_all: list[dict] = []
errors: list[str] = []
cancelados: list[dict] = []

# Store dos XMLs para download por nota (sig -> bytes/metadata)
st.session_state.setdefault('xml_store', {})
st.session_state.setdefault('nnf_to_sig', {})

# Acumuladores por NOTA (ICMSTot)
icms_total_all = 0.0
pis_total_all = 0.0
cofins_total_all = 0.0

if xml_files:
    # Mostra spinner enquanto processa uploads (XML/ZIP)
    spinner_placeholder.markdown(SPINNER_HTML, unsafe_allow_html=True)

    seen_xml_sigs: set[str] = set()
    dupes_ignored = 0
    xml_processed = 0

    for f in xml_files:
        try:
            b = f.read()
            # Deduplicação: evita processar o mesmo XML mais de uma vez
            if not f.name.lower().endswith(".zip"):
                sig = _xml_signature(b)
                if sig in seen_xml_sigs:
                    dupes_ignored += 1
                    continue
                seen_xml_sigs.add(sig)
                xml_processed += 1

            # Guardar XML para download individual (por assinatura/chave)
            try:
                root_tmp = ET.fromstring(b)
                nnf_tmp = _parse_nnf(root_tmp) or ""
                dh_tmp = _parse_date(root_tmp)
                chave_tmp = _extract_nfe_key(b)
            except Exception:
                nnf_tmp, dh_tmp, chave_tmp = "", None, ""
            st.session_state["xml_store"][sig] = {
                "bytes": b,
                "src": f.name,
                "Numero": nnf_tmp,
                "Data": dh_tmp,
                "chave": chave_tmp,
            }
            if nnf_tmp:
                st.session_state["nnf_to_sig"].setdefault(str(nnf_tmp), [])
                if sig not in st.session_state["nnf_to_sig"][str(nnf_tmp)]:
                    st.session_state["nnf_to_sig"][str(nnf_tmp)].append(sig)

                        # Guardar XML para download individual (por assinatura/chave)
                        try:
                            root_tmp = ET.fromstring(xb)
                            nnf_tmp = _parse_nnf(root_tmp) or ""
                            dh_tmp = _parse_date(root_tmp)
                            chave_tmp = _extract_nfe_key(xb)
                        except Exception:
                            nnf_tmp, dh_tmp, chave_tmp = "", None, ""
                        st.session_state["xml_store"][sig] = {
                            "bytes": xb,
                            "src": f"{f.name}:{xn}",
                            "Numero": nnf_tmp,
                            "Data": dh_tmp,
                            "chave": chave_tmp,
                        }
                        if nnf_tmp:
                            st.session_state["nnf_to_sig"].setdefault(str(nnf_tmp), [])
                            if sig not in st.session_state["nnf_to_sig"][str(nnf_tmp)]:
                                st.session_state["nnf_to_sig"][str(nnf_tmp)].append(sig)
            # Totais por NOTA (somente quando for XML direto)
            if not f.name.lower().endswith(".zip"):
                tot0 = _parse_tax_totals_from_xml(b)
                icms_total_all += tot0["vICMS"]
                pis_total_all += tot0["vPIS"]
                cofins_total_all += tot0["vCOFINS"]
            if f.name.lower().endswith(".zip"):
                with zipfile.ZipFile(io.BytesIO(b)) as z:
                    xml_names = sorted(set(n for n in z.namelist() if n.lower().endswith(".xml")))
                    if not xml_names:
                        errors.append(f"{f.name}: zip sem .xml")
                        continue
                    for xn in xml_names:
                        xb = z.read(xn)
                        sig = _xml_signature(xb)
                        if sig in seen_xml_sigs:
                            dupes_ignored += 1
                            continue
                        seen_xml_sigs.add(sig)
                        xml_processed += 1
                        tot = _parse_tax_totals_from_xml(xb)
                        icms_total_all += tot["vICMS"]
                        pis_total_all += tot["vPIS"]
                        cofins_total_all += tot["vCOFINS"]
                        rows = _parse_items_from_xml(xb, f"{f.name}:{xn}")
                        for rr in rows:
                            rr['xml_sig'] = sig
                        if not rows:
                            ce = _detect_cancel_event(xb)
                            if ce is not None:
                                ce["arquivo"] = f"{f.name}:{xn}"
                                cancelados.append(ce)
                                # evento de cancelamento não possui itens/IBSCBS
                                continue
                            errors.append(f"{f.name}:{xn}: não encontrei itens com IBSCBS")
                        rows_all.extend(rows)
            else:
                rows = _parse_items_from_xml(b, f.name)
                for rr in rows:
                    rr['xml_sig'] = sig
                if not rows:
                    ce = _detect_cancel_event(b)
                    if ce is not None:
                        ce["arquivo"] = f.name
                        cancelados.append(ce)
                    else:
                        errors.append(f"{f.name}: não encontrei itens com IBSCBS")
                rows_all.extend(rows)
        except Exception as e:
            errors.append(f"{f.name}: erro ao ler ({e})")

    # Remove spinner ao terminar
    spinner_placeholder.empty()

    if dupes_ignored:
        st.info(f"🔁 {dupes_ignored} XML(s) foram ignorados por duplicidade (mesma chave/conteúdo).")

df = pd.DataFrame(rows_all)

# Normaliza Data
if not df.empty:
    df["Data"] = pd.to_datetime(df["Data"], errors="coerce").dt.date

# ---------- KPIs ----------
def money(x):
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return "R$ 0,00"
    try:
        return "R$ {:,.2f}".format(float(x)).replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "R$ 0,00"

def pct(x):
    try:
        return f"{float(x):.1f}%"
    except Exception:
        return ""


def _fmt_money_br(x):
    try:
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return "0,00"
        return "{:,.2f}".format(float(x)).replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "0,00"

def _h(x):
    # escape for safe HTML rendering (keeps text)
    try:
        return html.escape("" if x is None else str(x))
    except Exception:
        return ""

def _clean_html(s: str) -> str:
    # Remove indentation that can turn HTML into a markdown code block
    return "\n".join(line.lstrip() for line in s.splitlines() if line.strip())

def _render_doc_table(df: pd.DataFrame, total_items: int | None = None):
    """
    Renderiza tabela premium (HTML) no estilo do print.
    """
    if df is None or df.empty:
        st.info("Nenhum item para exibir.")
        return

    total = total_items if total_items is not None else len(df)

    rows = []
    for _, r in df.iterrows():
        data = _h(r.get("Data", ""))
        numero = _h(r.get("Numero", ""))
        item = _h(r.get("Item/Serviço", ""))
        cclass = _h(r.get("cClassTrib", ""))
        valor = _fmt_money_br(r.get("Valor da operação", 0))
        vibs = _fmt_money_br(r.get("vIBS", 0))
        vcbs = _fmt_money_br(r.get("vCBS", 0))
        arquivo = _h(r.get("arquivo", ""))

        rows.append(f"""
<tr>
  <td class="col-date">{data}</td>
  <td class="col-num">{numero}</td>
  <td class="col-item">{item}</td>
  <td class="col-cclass"><span class="cclass-badge">{cclass}</span></td>
  <td class="col-money">{valor}</td>
  <td class="col-vibs">{vibs}</td>
  <td class="col-vcbs">{vcbs}</td>
  <td class="col-file" title="{arquivo}">{arquivo}</td>
</tr>
""")

    html_block = f"""
<div class="doc-table-wrap">
  <table class="doc-table">
    <thead>
      <tr>
        <th>DATA</th>
        <th>NÚMERO</th>
        <th>ITEM/SERVIÇO</th>
        <th>cClassTrib</th>
        <th>VALOR DA OPERAÇÃO</th>
        <th>vIBS</th>
        <th>vCBS</th>
        <th>ARQUIVO</th>
      </tr>
    </thead>
    <tbody>
      {''.join(rows)}
    </tbody>
  </table>
  <div class="doc-table-foot">Mostrando {len(df)} de {total} itens</div>
</div>
"""
    st.markdown(_clean_html(html_block), unsafe_allow_html=True)


# --- Totais (Somatório das bases do XML) ---
# Aqui os painéis mostram apenas a SOMA DAS BASES encontradas no XML (sem aplicar alíquota).
# As alíquotas exibidas são apenas informativas (fictícias), como você pediu.
ALIQUOTA_IBS_TEXTO = "0,10%"
ALIQUOTA_CBS_TEXTO = "0,90%"

base_ibs = float(df["Valor da operação"].fillna(0).sum()) if (not df.empty and "Valor da operação" in df.columns) else 0.0
base_cbs = float(df["Valor da operação"].fillna(0).sum()) if (not df.empty and "Valor da operação" in df.columns) else 0.0

# Totais exibidos nos cards = soma das bases
ibs_total = round(base_ibs, 2)
cbs_total = round(base_cbs, 2)
total_tributos = round(icms_total_all, 2)
# Créditos: Totais reais do XML (somatório de vIBS e vCBS)
creditos_ibs_total = round(float(df["vIBS"].fillna(0).sum()) if (not df.empty and "vIBS" in df.columns) else 0.0, 2)
creditos_cbs_total = round(float(df["vCBS"].fillna(0).sum()) if (not df.empty and "vCBS" in df.columns) else 0.0, 2)
# --- KPI clique (filtro via query param) ---
try:
    _qp = st.query_params.get("kpi", "all")
    # Streamlit pode devolver lista/tuple dependendo da versão
    if isinstance(_qp, (list, tuple)):
        selected_kpi = _qp[0] if _qp else "all"
    else:
        selected_kpi = _qp or "all"
except Exception:
    selected_kpi = "all"

selected_kpi = str(selected_kpi).lower().strip()
if selected_kpi not in ("all", "ibs", "cbs", "cred", "total"):
    selected_kpi = "all"


st.markdown(
    f"""
<div class="kpi-grid">
  <a class="kpi-link" href="?kpi=ibs">
    <div class="kpi kpi-ibs {'is-active' if selected_kpi=='ibs' else ''}">
      <div class="kpi-head">
        <div>
          <div class="label">IBS Total</div>
          <div class="pill">↗ Alíquota {ALIQUOTA_IBS_TEXTO}</div>
        </div>
        <div class="kpi-icon" aria-hidden="true">
          <svg viewBox="0 0 24 24" fill="none">
            <path d="M7 17V7" stroke="#2563eb" stroke-width="2" stroke-linecap="round"/>
            <path d="M7 17h10" stroke="#2563eb" stroke-width="2" stroke-linecap="round"/>
            <path d="M9 13l3-3 3 2 2-3" stroke="#2563eb" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
          </svg>
        </div>
      </div>
      <div class="value">{money(ibs_total)}</div>
      <div class="sub">Soma das bases IBS (XML)</div>
    </div>
  </a>

  <a class="kpi-link" href="?kpi=cbs">
    <div class="kpi kpi-cbs {'is-active' if selected_kpi=='cbs' else ''}">
      <div class="kpi-head">
        <div>
          <div class="label">CBS Total</div>
          <div class="pill">↗ Alíquota {ALIQUOTA_CBS_TEXTO}</div>
        </div>
        <div class="kpi-icon" aria-hidden="true">
          <svg viewBox="0 0 24 24" fill="none">
            <path d="M8 7h8M8 12h8M8 17h8" stroke="#16a34a" stroke-width="2" stroke-linecap="round"/>
            <path d="M6 5h12a2 2 0 0 1 2 2v10a2 2 0 0 1-2 2H6a2 2 0 0 1-2-2V7a2 2 0 0 1 2-2z" stroke="#16a34a" stroke-width="2"/>
          </svg>
        </div>
      </div>
      <div class="value">{money(cbs_total)}</div>
      <div class="sub">Soma das bases CBS (XML)</div>
    </div>
  </a>

  <a class="kpi-link" href="?kpi=cred">
    <div class="kpi kpi-cred {'is-active' if selected_kpi=='cred' else ''}">
      <div class="kpi-head">
        <div>
          <div class="label">Créditos</div>
          <div class="pill">↗ IBS + CBS</div>
        </div>
        <div class="kpi-icon" aria-hidden="true">
          <svg viewBox="0 0 24 24" fill="none">
            <path d="M7 12h10" stroke="#f59e0b" stroke-width="2" stroke-linecap="round"/>
            <path d="M12 7v10" stroke="#f59e0b" stroke-width="2" stroke-linecap="round"/>
            <path d="M6 6h12v12H6z" stroke="#f59e0b" stroke-width="2" opacity=".6"/>
          </svg>
        </div>
      </div>
      <div class="sub" style="margin-top:6px; font-weight:900; color:#f59e0b;">IBS: {money(creditos_ibs_total)}</div>
      <div class="sub" style="margin-top:6px; font-weight:900; color:#f59e0b;">CBS: {money(creditos_cbs_total)}</div>
      <div class="sub" style="margin-top:8px;">Totais extraídos de <b>vIBS</b> e <b>vCBS</b> (XML)</div>
    </div>
  </a>

  <a class="kpi-link" href="?kpi=total">
    <div class="kpi kpi-total {'is-active' if selected_kpi=='total' else ''}">
      <div class="kpi-head">
        <div>
          <div class="label">ICMS Total</div>
          <div class="pill">↗ ICMS</div>
        </div>
        <div class="kpi-icon" aria-hidden="true">
          <svg viewBox="0 0 24 24" fill="none">
            <path d="M7 8h10M7 12h10M7 16h10" stroke="#a855f7" stroke-width="2" stroke-linecap="round"/>
            <path d="M9 3h6v3H9z" stroke="#a855f7" stroke-width="2"/>
            <path d="M6 6h12v15H6z" stroke="#a855f7" stroke-width="2" opacity=".6"/>
          </svg>
        </div>
      </div>
      <div class="value">{money(total_tributos)}</div>
      <div class="sub">Somatório de ICMS (ICMSTot)</div>
    </div>
  </a>
</div>

<div style="margin-top:10px;">
  <a class="kpi-link" href="?kpi=all"><span class="pill">Limpar filtro</span></a>
</div>
""",
    unsafe_allow_html=True,
)
# Painéis (estilo Figma) — Totais por XML (ICMSTot)
c1, c2 = st.columns(2, gap="large")

pis_total = float(pis_total_all or 0.0)
cofins_total = float(cofins_total_all or 0.0)

with c1:
    st.markdown(
        f"""
<div class="card ibs-panel">
  <div class="panel-title">
    <div class="panel-left">
      <div class="icon" aria-hidden="true">
        <svg viewBox="0 0 24 24" fill="none">
          <path d="M4 18V6" stroke="#334155" stroke-width="1.7" stroke-linecap="round"/>
          <path d="M4 18h16" stroke="#334155" stroke-width="1.7" stroke-linecap="round"/>
          <path d="M8 14l3-3 3 2 4-5" stroke="#2563eb" stroke-width="1.9" stroke-linecap="round" stroke-linejoin="round"/>
        </svg>
      </div>
      <div>
        <h3>PIS - Total apurado</h3>
        <div class="hint">Somatório de vPIS (ICMSTot) de todos os XML</div>
      </div>
    </div>
    <span class="badge on">Ativo</span>
  </div>

  <div style="margin-top: 8px;">
    <div class="bar-label"><span>Total</span><span class="badge-money">{money(pis_total)}</span></div>
    <div class="bar-track"><div class="bar-fill ibs" style="width:100%"></div></div>
  </div>
</div>
""",
        unsafe_allow_html=True,
    )

with c2:
    st.markdown(
        f"""
<div class="card cbs-panel">
  <div class="panel-title">
    <div class="panel-left">
      <div class="icon" aria-hidden="true">
        <svg viewBox="0 0 24 24" fill="none">
          <path d="M4 18V6" stroke="#334155" stroke-width="1.7" stroke-linecap="round"/>
          <path d="M4 18h16" stroke="#334155" stroke-width="1.7" stroke-linecap="round"/>
          <path d="M8 14l3-3 3 2 4-5" stroke="#16a34a" stroke-width="1.9" stroke-linecap="round" stroke-linejoin="round"/>
        </svg>
      </div>
      <div>
        <h3>COFINS - Total apurado</h3>
        <div class="hint">Somatório de vCOFINS (ICMSTot) de todos os XML</div>
      </div>
    </div>
    <span class="badge on" style="background:#ecfdf3;border-color:#dcfce7;color:#166534;">Ativo</span>
  </div>

  <div style="margin-top: 8px;">
    <div class="bar-label"><span>Total</span><span class="badge-money">{money(cofins_total)}</span></div>
    <div class="bar-track"><div class="bar-fill cbs" style="width:100%"></div></div>
  </div>
</div>
""",
        unsafe_allow_html=True,
    )

st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

# Alerts
if cancelados:
    st.info(f"✅ {len(cancelados)} arquivo(s) são eventos de **cancelamento** e foram ignorados (não possuem itens/IBSCBS).")
    with st.expander("Ver cancelamentos detectados"):
        for c in cancelados[:20]:
            ch = c.get("chNFe", "") or "-"
            arq = c.get("arquivo", "") or "-"
            nprot = c.get("nProt", "") or "-"
            dh = (c.get("dhEvento", "") or "-")[:19].replace("T", " ")
            st.write(f"• {arq} | chNFe: {ch} | nProt: {nprot} | dhEvento: {dh}")
        if len(cancelados) > 20:
            st.caption(f"... e mais {len(cancelados)-20} cancelamentos")

if errors:
    st.warning("⚠️ Alguns arquivos não possuem bloco IBSCBS (ou não são NFe/NFC-e de itens):")
    for e in errors[:10]:
        st.write("•", e)
    if len(errors) > 10:
        st.caption(f"... e mais {len(errors)-10} itens")

# ---------- Filters + table ----------
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown("## Itens do Documento")
st.caption("Detalhamento dos itens extraídos do XML (inclui base vBC e valores de IBS/CBS quando presentes).")

if df.empty:
    st.info("Envie XML(s) para visualizar os itens aqui.")
    st.markdown("</div>", unsafe_allow_html=True)
    st.stop()

c1, c2, c3 = st.columns([1, 2, 1], gap="large")

with c1:
    min_d = df["Data"].min()
    max_d = df["Data"].max()
    # SEMPRE define "periodo" (evita NameError)
    periodo = st.date_input("Período", value=(min_d, max_d), min_value=min_d, max_value=max_d)

with c2:
    q = st.text_input("Buscar item", placeholder="Ex.: produto, serviço, descrição...")

with c3:
    classes = sorted([c for c in df["cClassTrib"].dropna().unique().tolist() if str(c).strip() != ""])
    pick = st.selectbox("cClassTrib", options=["(Todos)"] + classes, index=0)

df_view = df.copy()

# filtro de período (robusto)
if isinstance(periodo, (list, tuple)) and len(periodo) == 2:
    d1, d2 = periodo
    df_view["Data"] = pd.to_datetime(df_view["Data"], errors="coerce").dt.date
    df_view = df_view[(df_view["Data"] >= d1) & (df_view["Data"] <= d2)]

# busca
if q:
    qq = q.strip().lower()
    df_view = df_view[df_view["Item/Serviço"].fillna("").str.lower().str.contains(qq, na=False)]

# cClassTrib
if pick and pick != "(Todos)":
    df_view = df_view[df_view["cClassTrib"].astype(str) == str(pick)]


# filtro por KPI (clique nos cards)
if selected_kpi != "all":
    vibs = df_view["vIBS"].fillna(0) if "vIBS" in df_view.columns else None
    vcbs = df_view["vCBS"].fillna(0) if "vCBS" in df_view.columns else None

    if selected_kpi == "ibs" and vibs is not None:
        df_view = df_view[vibs != 0]
    elif selected_kpi == "cbs" and vcbs is not None:
        df_view = df_view[vcbs != 0]
    elif selected_kpi == "cred" and (vibs is not None and vcbs is not None):
        # créditos normalmente aparecem como valores negativos
        df_view = df_view[(vibs < 0) | (vcbs < 0)]
    elif selected_kpi == "total" and (vibs is not None and vcbs is not None):
        df_view = df_view[(vibs != 0) | (vcbs != 0)]


# ---------- Validação Premium IBS/CBS (retângulo) ----------
try:
    df_validado = aplicar_validacao_base_ibscbs(df_view)
    render_painel_validacao_premium(df_validado, key_prefix="ibscbs")
except Exception as _e:
    st.warning(f"Não foi possível renderizar a validação IBS/CBS: {_e}")


show_cols = ["Data", "Numero", "Item/Serviço", "cClassTrib", "Valor da operação", "vIBS", "vCBS", "arquivo", "Fonte do valor"]
show_cols = [c for c in show_cols if c in df_view.columns]

# ===== TABELA PREMIUM (igual vídeo) =====
st.markdown('<div class="table-wrap">', unsafe_allow_html=True)

_render_doc_table(df_view[show_cols], total_items=len(df_view))
st.markdown('<div class="table-download-spacer"></div>', unsafe_allow_html=True)
st.download_button(
    "Baixar CSV filtrado",
    data=df_view[show_cols].to_csv(index=False).encode("utf-8"),
    file_name="itens_filtrados.csv",
    mime="text/csv",
)

st.markdown('</div>', unsafe_allow_html=True)

# ---------- Generate planilha ----------
st.markdown('<div class="hr"></div>', unsafe_allow_html=True)
st.markdown("## Gerar planilha preenchida")

if template_bytes is None:
    st.error("Não encontrei **planilha_modelo.xlsx** na mesma pasta do app.py.")
else:
    if st.button("Gerar planilha", type="primary"):
        try:
            # 🔵 IBS
            show_spinner(tipo="ibs", titulo="Processando IBS…", subtitulo="Organizando bases", speed="1.6s")
            time.sleep(0.25)

            # 🟢 CBS
            show_spinner(tipo="cbs", titulo="Processando CBS…", subtitulo="Calculando valores", speed="1.4s")
            time.sleep(0.25)

            # 🟠 Créditos
            show_spinner(tipo="cred", titulo="Aplicando créditos…", subtitulo="Ajustando compensações", speed="1.2s")
            time.sleep(0.25)

            # 🟣 Total / exportação
            show_spinner(tipo="total", titulo="Gerando planilha…", subtitulo="Aplicando fórmulas e estilos", speed="1.0s")

            out_bytes = _append_to_workbook(template_bytes, df_view)

        except Exception as e:
            # Garante que o overlay não esconda o erro
            hide_spinner()
            st.error("Erro ao gerar a planilha. Veja os detalhes abaixo:")
            st.exception(e)
        else:
            hide_spinner()
            st.success("Planilha gerada! Abra no Excel para ver as fórmulas calculando.")

            st.download_button(
                "Baixar planilha_preenchida.xlsx",
                data=out_bytes,
                file_name="planilha_preenchida.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )